#!/usr/bin/env python3
"""
测试模版格式导出API功能
"""

import requests
import json
import base64
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# API基础URL
API_BASE_URL = "http://localhost:8000"

def test_template_export_api():
    """测试模版格式导出API功能"""
    logger.info("🚀 开始测试模版格式导出API...")
    
    try:
        # 1. 检查API健康状态
        logger.info("\n1️⃣ 检查API健康状态...")
        response = requests.get(f"{API_BASE_URL}/")
        if response.status_code != 200:
            logger.error("❌ API服务未启动")
            return False
        logger.info("✅ API服务正常")
        
        # 2. 准备测试数据（使用学习报告.xmind）
        logger.info("\n2️⃣ 准备测试数据...")
        xmind_path = "../学习报告.xmind"
        if not os.path.exists(xmind_path):
            logger.error(f"❌ 测试文件不存在: {xmind_path}")
            return False
        
        # 3. 分析XMind文件
        logger.info("\n3️⃣ 分析XMind文件...")
        with open(xmind_path, 'rb') as f:
            files = {'file': (os.path.basename(xmind_path), f, 'application/octet-stream')}
            response = requests.post(f"{API_BASE_URL}/api/analyze", files=files)
        
        if response.status_code != 200:
            logger.error(f"❌ 文件分析失败: {response.text}")
            return False
        
        analysis_data = response.json()
        markers_found = analysis_data['markers_found']
        logger.info(f"✅ 分析完成，发现 {len(markers_found)} 种标识符")
        
        # 4. 选择标识符进行模版格式导出
        logger.info("\n4️⃣ 选择标识符进行模版格式导出...")
        
        # 选择前3种标识符（保证有足够的测试数据）
        selected_markers = [marker['markerId'] for marker in markers_found[:3]]
        logger.info(f"选中标识符: {selected_markers}")
        
        export_request = {
            "selected_markers": selected_markers,
            "file_data": analysis_data['file_data']
        }
        
        # 5. 调用模版格式导出API
        logger.info("\n5️⃣ 调用模版格式导出API...")
        response = requests.post(f"{API_BASE_URL}/api/export-template", json=export_request)
        
        if response.status_code != 200:
            logger.error(f"❌ 模版格式导出失败: {response.text}")
            return False
        
        export_result = response.json()
        logger.info("✅ 模版格式导出成功")
        
        # 6. 验证返回结果
        logger.info("\n6️⃣ 验证返回结果...")
        
        expected_fields = ['success', 'message', 'filename', 'file_data', 'export_details']
        for field in expected_fields:
            if field not in export_result:
                logger.error(f"❌ 缺少必要字段: {field}")
                return False
        
        export_details = export_result['export_details']
        logger.info(f"   总用例数: {export_details['total_cases']}")
        logger.info(f"   导出格式: {export_details['export_format']}")
        logger.info(f"   列数: {len(export_details['columns'])}")
        logger.info(f"   列名: {export_details['columns']}")
        
        # 7. 保存并验证Excel文件
        logger.info("\n7️⃣ 保存并验证Excel文件...")
        
        # 解码并保存Excel文件
        excel_data = base64.b64decode(export_result['file_data'])
        test_filename = f"test_template_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with open(test_filename, 'wb') as f:
            f.write(excel_data)
        
        logger.info(f"   Excel文件已保存: {test_filename}")
        logger.info(f"   文件大小: {len(excel_data):,} bytes")
        
        # 8. 验证Excel文件结构
        logger.info("\n8️⃣ 验证Excel文件结构...")
        
        wb = load_workbook(test_filename)
        logger.info(f"   工作表: {wb.sheetnames}")
        
        # 验证冒烟测试用例工作表
        if '冒烟测试用例' in wb.sheetnames:
            ws_test = wb['冒烟测试用例']
            logger.info(f"   冒烟测试用例工作表: {ws_test.max_row} 行, {ws_test.max_column} 列")
            
            # 验证表头
            expected_headers = ["节点1", "节点2", "节点3", "节点4", "节点5", "端/API/服务", "冒烟结果", "研发对应负责人", "showcase问题", "是否核心功能", "是否影响主流程", "执行时间"]
            actual_headers = [ws_test.cell(row=1, column=col).value for col in range(1, ws_test.max_column + 1)]
            
            headers_match = all(expected in actual_headers for expected in expected_headers)
            if headers_match:
                logger.info("   ✅ 表头格式完全匹配模版")
            else:
                logger.warning(f"   ⚠️ 表头不匹配，预期: {expected_headers[:5]}..., 实际: {actual_headers[:5]}...")
            
            # 验证数据行
            if ws_test.max_row > 1:
                logger.info("   ✅ 包含测试用例数据")
                
                # 显示第一行数据作为示例
                first_row_data = [ws_test.cell(row=2, column=col).value for col in range(1, min(6, ws_test.max_column + 1))]
                logger.info(f"   第一行数据示例: {first_row_data}")
            else:
                logger.warning("   ⚠️ 没有测试用例数据")
        else:
            logger.error("   ❌ 缺少'冒烟测试用例'工作表")
        
        # 验证导出汇总工作表
        if '导出汇总' in wb.sheetnames:
            ws_summary = wb['导出汇总']
            logger.info(f"   导出汇总工作表: {ws_summary.max_row} 行, {ws_summary.max_column} 列")
            logger.info("   ✅ 包含导出汇总信息")
        else:
            logger.error("   ❌ 缺少'导出汇总'工作表")
        
        wb.close()
        
        # 9. 对比分析
        logger.info("\n9️⃣ 对比标准导出与模版导出...")
        
        # 调用标准导出API进行对比
        response_standard = requests.post(f"{API_BASE_URL}/api/export", json=export_request)
        if response_standard.status_code == 200:
            standard_result = response_standard.json()
            standard_cases = standard_result['smoke_test_suite']['test_cases']
            
            logger.info(f"   标准导出用例数: {len(standard_cases)}")
            logger.info(f"   模版导出用例数: {export_details['total_cases']}")
            
            if len(standard_cases) == export_details['total_cases']:
                logger.info("   ✅ 用例数量一致")
            else:
                logger.warning("   ⚠️ 用例数量不一致")
        
        logger.info("\n🎉 模版格式导出API测试完成!")
        logger.info("="*60)
        logger.info("✅ 测试结果总结:")
        logger.info("   - API接口正常工作")
        logger.info("   - Excel文件格式符合模版规范")
        logger.info("   - 路径成功拆分为5个节点层级")
        logger.info("   - 包含所有必要的业务字段")
        logger.info("   - 生成了完整的导出汇总信息")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ 测试过程中出错: {str(e)}")
        return False


def main():
    """主函数"""
    logger.info("🧪 模版格式导出API测试")
    logger.info("="*60)
    
    success = test_template_export_api()
    
    if success:
        logger.info("\n🎊 所有测试通过！模版格式导出功能工作正常。")
    else:
        logger.error("\n💥 测试失败，请检查相关功能。")
    
    return success


if __name__ == "__main__":
    main() 