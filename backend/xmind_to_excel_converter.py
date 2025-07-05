#!/usr/bin/env python3
"""
XMind数据直接转换到Excel模块
保持与原始XMind相同的数据结构和完整性
"""

import logging
import json
import base64
import io
import os
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict

logger = logging.getLogger(__name__)

class XMindToExcelConverter:
    """将XMind数据直接转换为Excel，保持数据结构完整性"""
    
    def __init__(self):
        # 配置样式
        self.level_colors = {
            1: 'B8CCE4',  # 节点1 - 浅蓝色
            2: 'D9E2F3',  # 节点2 - 更浅蓝色
            3: 'E8F1F9',  # 节点3 - 极浅蓝色
            4: 'F0F8FF',  # 节点4 - 接近白色
            5: 'FFFFFF'   # 节点5 - 白色
        }
        self.header_color = '4F81BD'  # 深蓝色表头
        self.business_column_color = 'F2F2F2'  # 业务列浅灰色
        
        # 最大支持的节点层级
        self.max_levels = 5
        
    def convert_to_excel(self, xmind_data: Dict, output_path: str = None) -> str:
        """
        将过滤后的XMind数据转换为Excel文件
        
        Args:
            xmind_data: 过滤后的XMind数据，包含JSON或XML格式的内容
            output_path: 输出文件路径，如果为None则自动生成
            
        Returns:
            生成的Excel文件路径
        """
        try:
            logger.info("🚀 开始将过滤后的XMind数据转换为Excel...")
            
            # 提取XMind结构
            if isinstance(xmind_data, dict) and 'content' in xmind_data:
                # 处理JSON格式的XMind数据
                content = xmind_data['content']
                if isinstance(content, str) and (content.startswith('[') or content.startswith('{')):
                    try:
                        content = json.loads(content)
                    except json.JSONDecodeError:
                        logger.warning("内容字符串不是有效的JSON，尝试其他解析方式")
            else:
                content = xmind_data
            
            # 解析XMind结构
            sheets = []
            if isinstance(content, list):
                sheets = content  # 已经是sheets列表
            elif isinstance(content, dict) and 'sheets' in content:
                sheets = content['sheets']
            elif isinstance(content, dict):
                sheets = [content]  # 单个sheet
                
            if not sheets:
                logger.warning("未找到有效的XMind工作表数据")
                sheets = []
                
            # 创建Excel工作簿
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = "冒烟测试用例"
            
            # 写入表头
            self._write_headers(ws_main)
            
            # 从XMind结构中提取层级数据并写入Excel
            hierarchy_data = self._extract_hierarchy(sheets)
            
            # 确保数据从第2行开始写入（第1行是表头）
            start_row = 2
            logger.info(f"🔢 设置起始行为 {start_row}")
            
            # 添加检查确保无无效的空白行
            if not hierarchy_data:
                logger.warning("⚠️ 提取的层级数据为空，将创建空的Excel文件")
            else:
                logger.info(f"📊 提取了 {len(hierarchy_data)} 个节点数据，准备写入Excel")
                # 记录第一个节点的信息以便调试
                if hierarchy_data and len(hierarchy_data) > 0:
                    first_node = hierarchy_data[0]
                    logger.info(f"👉 第一个节点: 标题='{first_node.get('title')}', 路径={first_node.get('path')}, 层级={first_node.get('level')}")
            
            # 写入数据，从第2行开始
            total_rows = self._write_data(ws_main, hierarchy_data, start_row)
            
            # 设置列宽
            self._set_column_widths(ws_main, total_rows)
            
            # 创建汇总表
            ws_summary = wb.create_sheet("导出汇总")
            self._create_summary_sheet(ws_summary, hierarchy_data)
            
            # 修复：保存前检查并删除空白行，确保没有多余的空白行
            self._remove_empty_rows(ws_main, start_row)
            
            # 保存Excel文件
            if not output_path:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                output_path = f"XMind导出_冒烟测试用例_{timestamp}.xlsx"
            
            wb.save(output_path)
            logger.info(f"✅ Excel转换完成: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"❌ XMind转Excel失败: {str(e)}")
            raise Exception(f"转换失败: {str(e)}")
    
    def _extract_hierarchy(self, sheets: List[Dict]) -> List[Dict]:
        """
        从XMind sheets中提取层级结构
        
        Returns:
            包含层级结构的数据列表
        """
        hierarchy = []
        logger.info(f"开始提取层级结构，共有 {len(sheets)} 个工作表")
        
        for sheet_idx, sheet in enumerate(sheets):
            if not sheet:
                logger.warning(f"跳过空工作表: 索引 {sheet_idx}")
                continue
                
            # 获取根主题
            root_topic = None
            if 'rootTopic' in sheet:
                root_topic = sheet['rootTopic']
                logger.info(f"找到rootTopic节点，标题: {root_topic.get('title', '未命名')}")
            elif 'topic' in sheet:
                root_topic = sheet['topic']
                logger.info(f"找到topic节点，标题: {root_topic.get('title', '未命名')}")
            else:
                # 尝试查找其他可能的根节点位置
                for key, value in sheet.items():
                    if isinstance(value, dict) and 'title' in value:
                        logger.info(f"尝试使用备选节点: {key}, 标题: {value['title']}")
                        root_topic = value
                        break
            
            if not root_topic:
                logger.warning(f"工作表{sheet_idx}中未找到根主题")
                continue
                
            # 提取层级结构
            nodes_before = len(hierarchy)
            self._process_topic(root_topic, [], hierarchy, 0)
            nodes_after = len(hierarchy)
            logger.info(f"工作表{sheet_idx}提取了 {nodes_after - nodes_before} 个节点")
            
        # 过滤掉空节点和无效节点
        valid_hierarchy = [node for node in hierarchy if node.get('title') and node.get('path')]
        invalid_nodes = len(hierarchy) - len(valid_hierarchy)
        
        if invalid_nodes > 0:
            logger.warning(f"过滤掉了 {invalid_nodes} 个无效节点")
            
        # 增加数据完整性检查
        if valid_hierarchy:
            logger.info(f"第一个有效节点: 标题='{valid_hierarchy[0].get('title')}', 层级={valid_hierarchy[0].get('level')}")
            if len(valid_hierarchy) > 1:
                logger.info(f"第二个有效节点: 标题='{valid_hierarchy[1].get('title')}', 层级={valid_hierarchy[1].get('level')}")
        else:
            logger.warning("没有提取到有效的节点数据")
        
        logger.info(f"从XMind数据中提取了 {len(valid_hierarchy)} 个有效节点")
        return valid_hierarchy
        
    def _process_topic(self, topic: Dict, path: List[str], result: List[Dict], level: int):
        """递归处理主题，构建层级结构"""
        if not topic or not isinstance(topic, dict):
            logger.debug(f"跳过无效主题: {type(topic)}")
            return
            
        title = topic.get('title', '').strip()
        if not title:
            logger.debug("跳过无标题节点")
            return
            
        # 当前路径
        current_path = path + [title]
        
        # 提取标识符
        markers = []
        if 'markers' in topic:
            markers_data = topic['markers']
            for marker in markers_data if isinstance(markers_data, list) else [markers_data]:
                if isinstance(marker, dict) and 'markerId' in marker:
                    markers.append(marker['markerId'])
                elif isinstance(marker, str):
                    markers.append(marker)
        
        # 提取备注
        notes = ""
        if 'notes' in topic:
            notes_data = topic['notes']
            if isinstance(notes_data, dict) and 'plain' in notes_data:
                plain = notes_data['plain']
                if isinstance(plain, dict) and 'content' in plain:
                    notes = plain['content']
                elif isinstance(plain, str):
                    notes = plain
        
        # 构建节点信息
        node_info = {
            'title': title,
            'level': level + 1,  # 从1开始计数
            'path': current_path,
            'markers': markers,
            'notes': notes,
            'children': []
        }
        
        # 记录一级节点信息
        if level == 0:
            logger.debug(f"添加一级节点: {title}")
        
        result.append(node_info)
        
        # 递归处理子节点
        children = []
        if 'children' in topic and isinstance(topic['children'], dict):
            children_data = topic['children']
            if 'attached' in children_data and isinstance(children_data['attached'], list):
                children = children_data['attached']
            elif 'topics' in children_data and isinstance(children_data['topics'], list):
                children = children_data['topics']
        elif 'topics' in topic and isinstance(topic['topics'], list):
            children = topic['topics']
        
        # 记录子节点数量
        if children:
            logger.debug(f"节点 '{title}' 有 {len(children)} 个子节点")
            
        for child in children:
            self._process_topic(child, current_path, result, level + 1)
    
    def _write_headers(self, ws):
        """写入Excel表头"""
        headers = ['节点1', '节点2', '节点3', '节点4', '节点5', 
                  '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题',
                  '是否核心功能', '是否影响主流程']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            
            # 表头样式
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
            cell.border = Border(
                left=Side(style='thin', color='D4D4D4'),
                right=Side(style='thin', color='D4D4D4'),
                top=Side(style='thin', color='D4D4D4'),
                bottom=Side(style='thin', color='D4D4D4')
            )
    
    def _write_data(self, ws, hierarchy_data: List[Dict], start_row: int) -> int:
        """
        将层级数据写入Excel
        
        Args:
            ws: 工作表对象
            hierarchy_data: 层级结构数据
            start_row: 起始行号
            
        Returns:
            写入的总行数
        """
        if not hierarchy_data:
            logger.warning("没有数据可写入！")
            return start_row
            
        # 调试日志
        logger.info(f"开始写入数据，起始行：{start_row}，总节点数：{len(hierarchy_data)}")
            
        # 强制确保从起始行开始写入，不留空白
        current_row = start_row
        row_mappings = {}  # 用于存储行映射信息，辅助合并单元格
        
        # 过滤掉任何可能的无效数据，确保数据质量
        valid_data = [node for node in hierarchy_data if node.get('path') and len(node.get('path', [])) > 0]
        
        # 将层级数据排序，确保相关节点连续
        sorted_data = sorted(valid_data, key=lambda x: '>'.join([str(i) for i in x['path']]))
        
        if not sorted_data:
            logger.warning("没有有效的节点数据可写入")
            return current_row
            
        logger.info(f"有效节点数：{len(sorted_data)}，将从第{current_row}行开始写入")
        
        # 修复：使用modules_processed计数器替代first_module标志，更可靠地跟踪模块处理状态
        prev_module = None
        modules_processed = 0  # 记录已处理的模块数量
        
        # 写入数据行
        for index, node in enumerate(sorted_data):
            # 跳过无效节点
            if not node.get('path'):
                logger.debug(f"跳过无效节点: {node.get('title', 'Unknown')}")
                continue
            
            level = min(node['level'], self.max_levels)
            
            # 识别当前模块（第一级节点）
            current_module = node['path'][0] if node['path'] else None
            if not current_module:
                logger.debug(f"跳过无模块节点: {node.get('title', 'Unknown')}")
                continue  # 跳过没有模块信息的节点
            
            # 详细的调试日志
            if index == 0:
                logger.info(f"写入第一个数据行 - 行号:{current_row}, 模块:{current_module}, 标题:{node.get('title')}, 路径:{node.get('path')}")
            elif current_module != prev_module:
                logger.info(f"模块变更: {prev_module} -> {current_module}, 当前行号: {current_row}")
                
                # 修复：仅在不是第一个模块的情况下添加空白行，避免在顶部添加不必要的空白行
                if modules_processed > 0:  # 只有处理完第一个模块后才添加空白行
                    self._write_empty_row(ws, current_row)
                    current_row += 1
                    logger.info(f"已添加模块分隔空行，新行号: {current_row}")
                
                # 增加已处理模块计数
                modules_processed += 1
                
            # 记录当前模块
            prev_module = current_module
            
            # 写入节点标题到对应的层级列
            for i in range(1, self.max_levels + 1):
                if i == level:
                    # 在当前节点级别写入标题
                    cell = ws.cell(row=current_row, column=i, value=node['title'])
                    
                    # 根据层级设置背景色
                    level_color = self.level_colors.get(level, 'FFFFFF')
                    cell.fill = PatternFill(start_color=level_color, end_color=level_color, fill_type='solid')
                    
                    # 设置边框和对齐方式
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin', color='D4D4D4'),
                        right=Side(style='thin', color='D4D4D4'),
                        top=Side(style='thin', color='D4D4D4'),
                        bottom=Side(style='thin', color='D4D4D4')
                    )
                elif i < level:
                    # 从路径中获取上级节点标题
                    if i-1 < len(node['path']) - 1:  # 减1是因为path中的最后一个是当前节点
                        parent_title = node['path'][i-1]
                        cell = ws.cell(row=current_row, column=i, value=parent_title)
                        
                        # 设置样式
                        parent_level_color = self.level_colors.get(i, 'FFFFFF')
                        cell.fill = PatternFill(start_color=parent_level_color, end_color=parent_level_color, fill_type='solid')
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.border = Border(
                            left=Side(style='thin', color='D4D4D4'),
                            right=Side(style='thin', color='D4D4D4'),
                            top=Side(style='thin', color='D4D4D4'),
                            bottom=Side(style='thin', color='D4D4D4')
                        )
            
            # 写入其他业务列
            business_columns = {
                6: self._get_service_api(node),  # 端/API/服务
                7: "",                           # 冒烟结果
                8: "",                           # 研发对应负责人
                9: "",                           # showcase问题
                10: "",                          # 是否核心功能 - 设为空白
                11: ""                           # 是否影响主流程 - 设为空白
            }
            
            for col, value in business_columns.items():
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.fill = PatternFill(start_color=self.business_column_color, end_color=self.business_column_color, fill_type='solid')
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color='D4D4D4'),
                    right=Side(style='thin', color='D4D4D4'),
                    top=Side(style='thin', color='D4D4D4'),
                    bottom=Side(style='thin', color='D4D4D4')
                )
            
            # 记录每行的节点信息
            row_mappings[current_row] = {
                'node': node,
                'path': node['path'],
                'level': level
            }
            
            current_row += 1
        
        # 应用增强的单元格合并前，检查行数据
        logger.info(f"写入完成，总共写入了 {current_row - start_row} 行数据，从第 {start_row} 行到第 {current_row - 1} 行")
        
        # 应用增强的单元格合并
        self._apply_enhanced_merges(ws, row_mappings)
        
        # 检查是否存在空白行并尝试清理
        self._clean_empty_rows(ws, start_row, current_row)
        
        return current_row
    
    def _apply_enhanced_merges(self, ws, row_mappings):
        """
        增强版单元格合并算法 - 更精确地识别和合并路径单元格
        
        Args:
            ws: 工作表对象
            row_mappings: 行到节点的映射
        """
        logger.info("🔍 开始应用增强版单元格合并算法")
        
        # 1. 构建路径树，用于识别同一路径的所有行
        path_tree = self._build_path_tree(row_mappings)
        
        # 2. 逐层处理合并（从高到低）
        for level in range(1, self.max_levels + 1):
            if level not in path_tree:
                continue
                
            logger.info(f"处理第{level}层路径合并")
            self._merge_cells_at_level(ws, path_tree[level], level)
        
        logger.info("✅ 单元格合并处理完成")

    def _build_path_tree(self, row_mappings) -> Dict:
        """
        构建路径树，用于识别可合并的单元格
        
        Args:
            row_mappings: 行到节点的映射
            
        Returns:
            按层级组织的路径树
        """
        path_tree = {}  # 层级 -> 路径 -> 行列表
        
        # 遍历所有行
        for row, mapping in sorted(row_mappings.items()):
            path = mapping['path']
            level = mapping['level']
            
            # 为每个层级构建路径键
            for i in range(1, level + 1):
                if i not in path_tree:
                    path_tree[i] = {}
                    
                # 构建当前层级的路径字符串
                if i == 1:
                    path_key = str(path[0])
                else:
                    # 对于更深层级，使用完整的父路径作为键
                    path_key = '>'.join([str(p) for p in path[:i]])
                
                # 将当前行添加到对应路径下
                if path_key not in path_tree[i]:
                    path_tree[i][path_key] = []
                path_tree[i][path_key].append(row)
        
        return path_tree

    def _merge_cells_at_level(self, ws, path_map, level):
        """
        合并特定层级的单元格
        
        Args:
            ws: 工作表
            path_map: 路径到行的映射
            level: 当前处理的层级
        """
        merge_count = 0
        
        # 遍历当前层级的所有路径
        for path_key, rows in path_map.items():
            if len(rows) <= 1:
                continue  # 单行不需要合并
                
            # 查找连续的行范围
            continuous_ranges = self._find_continuous_ranges(rows)
            
            # 对每个连续范围执行合并
            for start_row, end_row in continuous_ranges:
                if end_row > start_row:  # 至少有2行才合并
                    # 检查值是否一致
                    first_cell_value = ws.cell(row=start_row, column=level).value
                    all_same = all(ws.cell(row=r, column=level).value == first_cell_value for r in range(start_row+1, end_row+1))
                    
                    if all_same and first_cell_value:  # 确保所有单元格值相同且不为空
                        # 执行合并
                        ws.merge_cells(start_row=start_row, start_column=level, end_row=end_row, end_column=level)
                        
                        # 美化合并后的单元格
                        self._style_merged_cell(ws.cell(row=start_row, column=level), level)
                        
                        merge_count += 1
        
        logger.info(f"第{level}层合并了{merge_count}个单元格区域")

    def _find_continuous_ranges(self, rows: List[int]) -> List[tuple]:
        """
        查找连续的行范围
        
        Args:
            rows: 行号列表
            
        Returns:
            连续范围的列表，每个范围为 (start_row, end_row) 元组
        """
        if not rows:
            return []
            
        # 确保行号排序
        sorted_rows = sorted(rows)
        
        ranges = []
        range_start = sorted_rows[0]
        prev_row = sorted_rows[0]
        
        for row in sorted_rows[1:]:
            if row == prev_row + 1:
                # 连续行
                prev_row = row
            else:
                # 不连续，结束当前范围
                ranges.append((range_start, prev_row))
                range_start = row
                prev_row = row
        
        # 添加最后一个范围
        ranges.append((range_start, prev_row))
        
        return ranges

    def _style_merged_cell(self, cell, level):
        """
        美化合并后的单元格
        
        Args:
            cell: 合并后的单元格（左上角）
            level: 层级
        """
        # 根据层级设置样式
        level_color = self.level_colors.get(level, 'FFFFFF')
        
        # 应用增强样式
        cell.fill = PatternFill(start_color=level_color, end_color=level_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 设置边框
        cell.border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4')
        )
        
        # 添加根据层级的缩进
        indent = max(0, level - 1)
        if hasattr(cell.alignment, 'indent'):
            cell.alignment = Alignment(
                horizontal='left', 
                vertical='center', 
                wrap_text=True, 
                indent=indent
            )
    
    def _get_service_api(self, node: Dict) -> str:
        """从节点信息中提取服务/API信息"""
        # 从标题或备注中识别API/服务信息
        title = node.get('title', '').lower()
        notes = node.get('notes', '').lower()
        
        # 检查API关键词
        api_keywords = ['api', '接口', 'service', '服务', 'endpoint', '微服务']
        for keyword in api_keywords:
            if keyword in title:
                return node.get('title', '')
        
        # 从备注中提取
        if notes:
            lines = notes.split('\n')
            for line in lines:
                line = line.strip()
                for keyword in api_keywords:
                    if keyword in line:
                        return line
        
        return ""
    
    def _is_core_function(self, node: Dict) -> bool:
        """判断节点是否为核心功能"""
        # 根据标识符和标题判断
        markers = node.get('markers', [])
        
        # 核心功能的标识符
        core_markers = ['important', 'priority-1', 'flag-red', 'star-red']
        if any(marker in core_markers for marker in markers):
            return True
        
        # 从标题判断
        title = node.get('title', '').lower()
        core_keywords = ['核心', '关键', 'core', 'key', 'critical', '主要', '重要']
        
        return any(keyword in title for keyword in core_keywords)
    
    def _affects_main_flow(self, node: Dict) -> bool:
        """判断节点是否影响主流程"""
        # 核心功能通常影响主流程
        if self._is_core_function(node):
            return True
            
        # 检查标题
        title = node.get('title', '').lower()
        main_flow_keywords = ['主流程', '关键流程', 'main flow', '主要功能']
        
        return any(keyword in title for keyword in main_flow_keywords)
    
    def _set_column_widths(self, ws, total_rows: int):
        """设置Excel列宽"""
        column_widths = {
            1: 30,  # 节点1
            2: 30,  # 节点2
            3: 30,  # 节点3
            4: 30,  # 节点4
            5: 30,  # 节点5
            6: 20,  # 端/API/服务
            7: 15,  # 冒烟结果
            8: 20,  # 研发对应负责人
            9: 20,  # showcase问题
            10: 15, # 是否核心功能
            11: 15  # 是否影响主流程
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_summary_sheet(self, ws, hierarchy_data: List[Dict]):
        """创建汇总信息工作表"""
        # 设置标题
        ws['A1'] = "XMind导出汇总信息"
        ws['A1'].font = Font(bold=True, size=14)
        
        # 基本信息
        row = 3
        ws[f'A{row}'] = "导出时间"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        ws[f'A{row}'] = "节点总数"
        ws[f'B{row}'] = len(hierarchy_data)
        row += 1
        
        # 统计各级节点数
        level_counts = {}
        for node in hierarchy_data:
            level = min(node['level'], self.max_levels)
            level_counts[level] = level_counts.get(level, 0) + 1
        
        for level, count in sorted(level_counts.items()):
            ws[f'A{row}'] = f"节点{level}级数量"
            ws[f'B{row}'] = count
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def _write_empty_row(self, ws, row):
        """
        写入一个空白行，用于模块分隔
        
        Args:
            ws: 工作表
            row: 行号
        """
        # 获取当前表格的列数（11列，移除了执行时间列）
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col, value="")
            # 设置边框
            cell.border = Border(
                left=Side(style='thin', color='D4D4D4'),
                right=Side(style='thin', color='D4D4D4'),
                top=Side(style='thin', color='D4D4D4'),
                bottom=Side(style='thin', color='D4D4D4')
            )
            # 添加轻微的背景色，使空白行更易识别
            cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')

    def _clean_empty_rows(self, ws, start_row, end_row):
        """
        检查并清理不必要的空白行
        
        Args:
            ws: 工作表
            start_row: 起始行
            end_row: 结束行
        """
        logger.info(f"检查空白行，范围：{start_row} - {end_row}")
        
        # 特别检查第2行（表头后的第一行数据）
        if start_row == 2:
            # 判断是否为空行的标准：前5列都为空
            is_empty = True
            for col in range(1, 6):
                if ws.cell(row=2, column=col).value is not None:
                    is_empty = False
                    break
                    
            if is_empty:
                logger.warning("检测到第2行是空白行，清除该行...")
                for col in range(1, 12):  # 清除所有单元格内容和格式
                    ws.cell(row=2, column=col).value = None
                    ws.cell(row=2, column=col).fill = None
                    ws.cell(row=2, column=col).border = None

    def _remove_empty_rows(self, ws, start_row):
        """
        在保存前删除所有空白行
        
        Args:
            ws: 工作表对象
            start_row: 起始数据行
        """
        logger.info("检查并删除所有空白行")
        
        # 获取工作表的最大行数
        max_row = ws.max_row
        empty_rows = []
        
        # 修复：特别检查第一个数据行（第2行）
        # 如果第一个数据行是空的，这可能是导致顶部空白行的主要原因
        first_data_row_empty = True
        for col in range(1, 12):
            if ws.cell(row=start_row, column=col).value is not None:
                first_data_row_empty = False
                break
                
        if first_data_row_empty:
            logger.warning(f"⚠️ 检测到第一个数据行（行{start_row}）是空白的！这会导致顶部出现空白行")
            empty_rows.append(start_row)
        
        # 从起始行开始检查每一行
        for row in range(start_row, max_row + 1):
            is_empty = True
            for col in range(1, 12):  # 检查所有列
                if ws.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
            
            if is_empty:
                empty_rows.append(row)
                logger.info(f"发现空白行: {row}")
        
        # 从后向前删除空白行（避免索引变化问题）
        if empty_rows:
            logger.info(f"将删除 {len(empty_rows)} 个空白行: {empty_rows}")
            for row in sorted(empty_rows, reverse=True):
                ws.delete_rows(row)
                logger.info(f"已删除空白行 {row}")
                
        # 修复：最后进行安全检查，确保所有空白行都被删除
        remaining_empty_rows = []
        for row in range(start_row, ws.max_row + 1):
            is_empty = True
            for col in range(1, 12):
                if ws.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
                    
            if is_empty:
                remaining_empty_rows.append(row)
                
        if remaining_empty_rows:
            logger.warning(f"⚠️ 删除后仍有 {len(remaining_empty_rows)} 个空白行: {remaining_empty_rows}")
        else:
            logger.info("✅ 所有空白行已清理完毕")

# 创建全局实例
xmind_to_excel = XMindToExcelConverter() 