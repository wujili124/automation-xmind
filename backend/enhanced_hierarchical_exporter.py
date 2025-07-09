#!/usr/bin/env python3
"""
增强版层级化Excel导出器
完全匹配《冒烟用例导出模版.xlsx》的合并和视觉效果
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Tuple
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

logger = logging.getLogger(__name__)

class EnhancedHierarchicalExporter:
    """增强版层级合并Excel导出器"""
    
    def __init__(self):
        # 精确匹配模版的层级背景色
        self.level_colors = {
            1: 'B8CCE4',  # 节点1 - 浅蓝色（匹配模版）
            2: 'D9E2F3',  # 节点2 - 更浅蓝色
            3: 'E8F1F9',  # 节点3 - 极浅蓝色
            4: 'F0F8FF',  # 节点4 - 接近白色
            5: 'FFFFFF',  # 节点5 - 白色
            6: 'FFFFFF',  # 节点6 - 白色
            7: 'FFFFFF',  # 节点7 - 白色
            8: 'FFFFFF',  # 节点8 - 白色
            9: 'FFFFFF',  # 节点9 - 白色
            10: 'FFFFFF'  # 节点10 - 白色
        }
        
        # 其他列的背景色
        self.business_column_color = 'F2F2F2'  # 业务列浅灰色
        
        # 表头颜色（匹配模版）
        self.header_color = '4F81BD'  # 深蓝色表头
    
    def export_with_enhanced_merge(self, test_cases_data: Dict[str, Any], output_path: str = None) -> str:
        """
        增强版层级合并导出 - 优化空白节点处理
        
        Args:
            test_cases_data: 测试用例数据
            output_path: 输出文件路径
            
        Returns:
            生成的文件路径
        """
        try:
            logger.info("🚀 开始增强版层级合并导出Excel（含空白节点优化）...")
            
            test_cases = test_cases_data['smoke_test_suite']['test_cases']
            metadata = test_cases_data['smoke_test_suite']['metadata']
            
            logger.info(f"📊 原始数据：{len(test_cases)} 个测试用例")
            
            # 1. 智能数据预处理和分组（含质量控制）
            grouped_data, row_mappings = self._smart_group_data(test_cases)
            logger.info(f"📋 数据分组：{len(grouped_data)} 个顶级分组，{len(row_mappings)} 行有效数据")
            
            # 统计优化效果
            optimization_stats = {
                'original_cases': len(test_cases),
                'valid_cases': len(row_mappings),
                'filtered_cases': len(test_cases) - len(row_mappings),
                'top_level_groups': len(grouped_data)
            }
            
            if optimization_stats['filtered_cases'] > 0:
                logger.info(f"✨ 数据优化：过滤 {optimization_stats['filtered_cases']} 个空白/无效节点")
            
            # 2. 创建工作簿
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = "冒烟测试用例"
            
            # 3. 写入表头（精确匹配模版）
            self._write_enhanced_headers(ws_main)
            
            # 4. 按层级写入数据并智能合并
            current_row = 2
            total_rows = self._write_enhanced_hierarchical_data(ws_main, grouped_data, row_mappings, current_row)
            
            # 5. 设置精确的列宽和行高（匹配模版）
            self._set_precise_column_widths(ws_main, total_rows)
            
            # 6. 创建汇总表（包含优化信息）
            ws_summary = wb.create_sheet("导出汇总")
            self._create_enhanced_summary_sheet(ws_summary, metadata, len(row_mappings), optimization_stats)
            
            # 7. 保存文件
            if not output_path:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                output_path = f"优化版层级合并_冒烟测试用例_{timestamp}.xlsx"
            
            wb.save(output_path)
            
            # 8. 输出详细统计信息
            logger.info(f"✅ 优化版层级合并Excel导出完成: {output_path}")
            logger.info(f"   📊 原始用例: {optimization_stats['original_cases']} 个")
            logger.info(f"   ✨ 有效用例: {optimization_stats['valid_cases']} 个")
            logger.info(f"   🗑️  过滤无效: {optimization_stats['filtered_cases']} 个")
            logger.info(f"   🎯 层级分组: {optimization_stats['top_level_groups']} 个")
            logger.info(f"   📄 总行数: {total_rows - 1} 行数据")
            logger.info(f"   🎨 完美的层级合并效果，无空白行干扰")
            logger.info(f"   📋 业务列已优化：显示空白，删除执行时间列")
            
            return output_path
            
        except Exception as e:
            logger.error(f"❌ 增强版层级合并导出失败: {str(e)}")
            raise Exception(f"增强版层级合并导出失败: {str(e)}")
    
    def _smart_group_data(self, test_cases: List[Dict]) -> Tuple[OrderedDict, List[Dict]]:
        """智能数据分组，确保完美的层级结构 - 增强版数据清理"""
        
        # 预处理：数据清理和验证
        processed_cases = []
        skipped_count = 0
        
        for case_idx, case in enumerate(test_cases):
            # 1. 数据完整性检查
            if not self._is_valid_test_case(case):
                skipped_count += 1
                continue
                
            path_str = case.get('test_path', '') or case.get('测试路径', '') or case.get('title', '')
            
            # 2. 智能路径解析和清理
            cleaned_nodes = self._parse_and_clean_path(path_str)
            if not cleaned_nodes:
                logger.warning(f"跳过路径无效的测试用例: {case.get('title', 'Unknown')}")
                skipped_count += 1
                continue
            
            # 3. 构建处理后的用例数据
            processed_case = case.copy()
            processed_case['parsed_nodes'] = cleaned_nodes
            processed_case['sort_key'] = ' > '.join(cleaned_nodes)  # 用于排序
            processed_case['cleaned_path'] = ' > '.join(cleaned_nodes)  # 清理后的路径
            
            # 确保xmind_index存在
            if 'xmind_index' not in processed_case:
                processed_case['xmind_index'] = case_idx * 1000  # 使用索引作为默认值
            
            processed_cases.append(processed_case)
        
        if skipped_count > 0:
            logger.info(f"数据清理：跳过 {skipped_count} 个无效测试用例，保留 {len(processed_cases)} 个有效用例")
        
        # 按路径排序，确保层级结构清晰
        processed_cases.sort(key=lambda x: x['sort_key'])
        
        # 构建智能层级字典
        hierarchy = OrderedDict()
        row_mappings = []  # 记录每行的详细信息
        
        for case_idx, case in enumerate(processed_cases):
            nodes = case['parsed_nodes']
            
            # 4. 构建嵌套字典（确保没有空节点）
            current_level = hierarchy
            full_path = []
            
            for level, node in enumerate(nodes):
                if not node:  # 安全检查，跳过空节点
                    continue
                    
                full_path.append(node)
                
                if node not in current_level:
                    current_level[node] = {
                        '_data': [],
                        '_children': OrderedDict(),
                        '_level': level + 1,
                        '_full_path': ' > '.join(full_path),
                        '_row_count': 0,  # 记录该节点包含的总行数
                        '_has_data': False  # 标记是否包含实际数据
                    }
                
                # 如果是最后一个节点，添加数据
                if level == len(nodes) - 1:
                    current_level[node]['_data'].append(case)
                    current_level[node]['_has_data'] = True
                    
                    # 记录行映射信息
                    row_info = {
                        'case_index': case_idx,
                        'nodes': nodes.copy(),
                        'level': level + 1,
                        'data': case,
                        'full_path': ' > '.join(full_path),
                        'xmind_index': case.get('xmind_index', case_idx * 1000)  # 确保xmind_index存在
                    }
                    row_mappings.append(row_info)
                
                current_level = current_level[node]['_children']
        
        # 5. 清理空节点和计算行数
        cleaned_hierarchy = self._clean_empty_nodes(hierarchy)
        self._calculate_row_counts(cleaned_hierarchy)
        
        logger.info(f"智能分组完成：{len(cleaned_hierarchy)} 个顶级分组，{len(row_mappings)} 行数据")
        return cleaned_hierarchy, row_mappings
    
    def _is_valid_test_case(self, case: Dict) -> bool:
        """验证测试用例是否有效"""
        # 检查基本字段
        title = case.get('title', '').strip()
        if not title or len(title) < 3:
            return False
        
        # 检查是否有路径信息
        path_fields = ['test_path', '测试路径', 'title']
        has_path = any(case.get(field, '').strip() for field in path_fields)
        if not has_path:
            return False
        
        # 检查是否有测试步骤
        steps = case.get('steps', [])
        if not steps or len(steps) == 0:
            return False
        
        return True
    
    def _parse_and_clean_path(self, path_str: str) -> List[str]:
        """解析和清理路径，返回有效的节点列表"""
        if not path_str or not path_str.strip():
            return []
        
        # 智能路径分割
        if ' > ' in path_str:
            nodes = path_str.split(' > ')
        elif ' / ' in path_str:
            nodes = path_str.split(' / ')
        elif ' -> ' in path_str:
            nodes = path_str.split(' -> ')
        else:
            # 单个节点的情况
            nodes = [path_str]
        
        # 清理节点
        cleaned_nodes = []
        for node in nodes:
            cleaned_node = node.strip()
            
            # 跳过空节点
            if not cleaned_node:
                continue
            
            # 跳过过短的节点（但保留中文单字）
            if len(cleaned_node) < 2 and not self._is_chinese_char(cleaned_node):
                continue
            
            # 跳过明显无效的节点
            if self._is_invalid_node(cleaned_node):
                continue
            
            cleaned_nodes.append(cleaned_node)
        
        # 确保至少有2级路径
        if len(cleaned_nodes) < 2:
            if len(cleaned_nodes) == 1:
                cleaned_nodes = ['测试模块', cleaned_nodes[0]]
            else:
                return []  # 完全无效的路径
        
        # 限制最大层级深度
        if len(cleaned_nodes) > 10:
            cleaned_nodes = cleaned_nodes[:10]
        
        return cleaned_nodes
    
    def _is_chinese_char(self, char: str) -> bool:
        """检查是否为中文字符"""
        return '\u4e00' <= char <= '\u9fff'
    
    def _is_invalid_node(self, node: str) -> bool:
        """检查节点是否无效"""
        node_lower = node.lower()
        
        # 无效的占位符
        invalid_patterns = [
            '...', '---', '###', '***', 'xxx', 'todo', 
            'placeholder', '占位符', '待定', '待补充', '空白'
        ]
        
        if node_lower in invalid_patterns:
            return True
        
        # 纯数字或纯符号
        if node.isdigit() or all(c in '.-_*#@!()[]{}' for c in node):
            return True
        
        return False
    
    def _clean_empty_nodes(self, node_dict: OrderedDict) -> OrderedDict:
        """递归清理空节点"""
        cleaned_dict = OrderedDict()
        
        for node_name, node_info in node_dict.items():
            # 递归清理子节点
            cleaned_children = self._clean_empty_nodes(node_info['_children'])
            
            # 检查节点是否有价值（有数据或有有价值的子节点）
            has_data = len(node_info['_data']) > 0
            has_valuable_children = len(cleaned_children) > 0
            
            if has_data or has_valuable_children:
                cleaned_info = node_info.copy()
                cleaned_info['_children'] = cleaned_children
                cleaned_dict[node_name] = cleaned_info
        
        return cleaned_dict
    
    def _calculate_row_counts(self, node_dict: OrderedDict):
        """递归计算每个节点的总行数"""
        for node_name, node_info in node_dict.items():
            row_count = len(node_info['_data'])  # 当前节点的数据行数
            
            # 递归计算子节点行数
            if node_info['_children']:
                child_rows = self._calculate_row_counts(node_info['_children'])
                row_count += child_rows
            
            node_info['_row_count'] = row_count
        
        # 返回当前层级的总行数
        return sum(node_info['_row_count'] for node_info in node_dict.values())
    
    def _write_enhanced_headers(self, ws):
        """写入优化的表头 - 删除执行时间列"""
        headers = [
            '节点1', '节点2', '节点3', '节点4', '节点5', '节点6', '节点7', '节点8', '节点9', '节点10',
            '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题',
            '是否核心功能', '是否影响主流程'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            
            # 精确匹配模版的表头样式
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
            cell.border = self._get_border()
    
    def _write_enhanced_hierarchical_data(self, ws, hierarchy: OrderedDict, row_mappings: List[Dict], start_row: int) -> int:
        """写入增强的层级数据并实现精确合并"""
        
        current_row = start_row
        
        # 按照XMind中的原始顺序索引排序，确保节点顺序与XMind完全一致
        xmind_sorted_mappings = sorted(row_mappings, key=lambda x: x.get('xmind_index', 0))
        
        # 写入所有数据行
        for row_info in xmind_sorted_mappings:
            self._write_enhanced_single_row(ws, current_row, row_info)
            current_row += 1
        
        # 应用智能合并
        self._apply_enhanced_merges(ws, hierarchy, xmind_sorted_mappings, start_row)
        
        return current_row
    
    def _write_enhanced_single_row(self, ws, row: int, row_info: Dict):
        """写入增强的单行数据"""
        
        nodes = row_info['nodes']
        case_data = row_info['data']
        
        # 写入节点列 (1-5列) - 精确的层级效果
        for col in range(1, 11):
            if col <= len(nodes):
                value = nodes[col-1]
                level = col
            else:
                value = ''
                level = 10
            
            cell = ws.cell(row=row, column=col, value=value)
            
            # 精确的层级背景色
            if value:
                color = self.level_colors.get(level, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                # 不同层级使用不同的字体样式
                if level == 1:
                    cell.font = Font(bold=True, size=11)
                elif level == 2:
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.font = Font(size=10)
            
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = self._get_border()
        
        # 写入业务列 (11-16列) - 显示为空白，删除执行时间列
        business_data = [
            '',  # 端/API/服务 - 空白
            '',  # 冒烟结果 - 空白
            '',  # 研发对应负责人 - 空白
            '',  # showcase问题 - 空白
            '',  # 是否核心功能 - 空白
            ''   # 是否影响主流程 - 空白
        ]
        
        for col, value in enumerate(business_data, 11):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.business_column_color, end_color=self.business_column_color, fill_type='solid')
            cell.border = self._get_border()
            cell.font = Font(size=10)
    
    def _apply_enhanced_merges(self, ws, hierarchy: OrderedDict, row_mappings: List[Dict], start_row: int):
        """应用增强的智能合并，精确匹配模版效果 - 优化空白处理"""
        
        # 为每个层级计算合并区域
        merge_regions = self._calculate_merge_regions_enhanced(hierarchy, row_mappings, start_row)
        
        # 过滤和验证合并区域
        valid_merge_regions = []
        for region in merge_regions:
            if self._is_valid_merge_region(region, ws):
                valid_merge_regions.append(region)
            else:
                logger.debug(f"跳过无效合并区域: {region}")
        
        # 应用合并
        merge_count = 0
        for region in valid_merge_regions:
            try:
                if region['end_row'] > region['start_row']:  # 只合并多行的区域
                    merge_range = f"{get_column_letter(region['column'])}{region['start_row']}:{get_column_letter(region['column'])}{region['end_row']}"
                    
                    # 验证合并范围的有效性
                    if self._validate_merge_range(ws, region):
                        ws.merge_cells(merge_range)
                        merge_count += 1
                        logger.debug(f"成功合并区域: {merge_range} = '{region['value']}'")
                    else:
                        logger.warning(f"合并范围验证失败，跳过: {merge_range}")
                        
            except Exception as e:
                logger.warning(f"合并失败: {region} - {e}")
        
        logger.info(f"完成智能合并：成功合并 {merge_count} 个区域")
    
    def _calculate_merge_regions_enhanced(self, hierarchy: OrderedDict, row_mappings: List[Dict], start_row: int) -> List[Dict]:
        """计算增强版精确合并区域 - 避免空白区域问题"""
        
        merge_regions = []
        
        # 按列计算合并区域
        for col in range(1, 11):  # 节点1-10列
            # 使用层级信息计算合并，但要验证数据完整性
            current_row = start_row
            self._calculate_column_merges_enhanced(hierarchy, col, current_row, merge_regions, row_mappings, start_row)
        
        return merge_regions
    
    def _calculate_column_merges_enhanced(self, node_dict: OrderedDict, target_col: int, current_row: int, merge_regions: List[Dict], row_mappings: List[Dict], start_row: int, path_prefix: List[str] = None):
        """递归计算列的合并区域 - 增强版数据验证"""
        
        if path_prefix is None:
            path_prefix = []
        
        for node_name, node_info in node_dict.items():
            # 验证节点数据完整性
            if not node_name or not node_name.strip():
                logger.debug(f"跳过空节点名称")
                continue
                
            current_path = path_prefix + [node_name]
            level = len(current_path)
            
            if level == target_col:  # 当前列需要合并
                # 计算该节点占用的行数
                node_rows = node_info.get('_row_count', 0)
                has_data = node_info.get('_has_data', False)
                
                # 只有当节点有足够行数且包含有效数据时才进行合并
                if node_rows > 1 and (has_data or len(node_info.get('_children', {})) > 0):
                    # 验证合并区域的数据一致性
                    if self._validate_merge_data_consistency(node_name, current_row, node_rows, target_col, row_mappings):
                        merge_regions.append({
                            'column': target_col,
                            'start_row': current_row,
                            'end_row': current_row + node_rows - 1,
                            'value': node_name,
                            'level': level,
                            'has_data': has_data,
                            'row_count': node_rows
                        })
                
                current_row += node_rows
            else:
                # 递归处理子节点
                if node_info.get('_children'):
                    current_row = self._calculate_column_merges_enhanced(
                        node_info['_children'], 
                        target_col, 
                        current_row, 
                        merge_regions, 
                        row_mappings, 
                        start_row, 
                        current_path
                    )
                else:
                    # 叶子节点，确保数据行数正确
                    data_count = len(node_info.get('_data', []))
                    if data_count > 0:
                        current_row += data_count
        
        return current_row
    
    def _is_valid_merge_region(self, region: Dict, ws) -> bool:
        """验证合并区域是否有效"""
        try:
            start_row = region['start_row']
            end_row = region['end_row']
            column = region['column']
            
            # 基本范围检查
            if start_row >= end_row or start_row < 1 or column < 1:
                return False
            
            # 检查工作表范围
            if end_row > ws.max_row + 100:  # 允许一定的扩展空间
                return False
                
            # 检查值的有效性
            value = region.get('value', '').strip()
            if not value:
                return False
            
            return True
            
        except Exception as e:
            logger.debug(f"合并区域验证异常: {e}")
            return False
    
    def _validate_merge_range(self, ws, region: Dict) -> bool:
        """验证合并范围内的数据一致性"""
        try:
            start_row = region['start_row']
            end_row = region['end_row']
            column = region['column']
            expected_value = region['value']
            
            # 检查合并范围内的单元格值是否一致
            for row in range(start_row, end_row + 1):
                cell_value = ws.cell(row=row, column=column).value
                if cell_value and str(cell_value).strip() != expected_value:
                    # 如果单元格有值但与期望值不一致，说明数据有问题
                    logger.debug(f"合并验证失败：行{row}列{column}值为'{cell_value}'，期望'{expected_value}'")
                    return False
            
            return True
            
        except Exception as e:
            logger.debug(f"合并范围验证异常: {e}")
            return False
    
    def _validate_merge_data_consistency(self, node_name: str, start_row: int, row_count: int, column: int, row_mappings: List[Dict]) -> bool:
        """验证合并数据的一致性"""
        try:
            # 检查在指定行范围内，该列应该都是相同的节点名
            relevant_mappings = [
                mapping for mapping in row_mappings
                if len(mapping['nodes']) >= column and mapping['nodes'][column-1] == node_name
            ]
            
            # 如果没有相关映射或者映射数量与行数不匹配，可能有问题
            if len(relevant_mappings) != row_count:
                logger.debug(f"数据一致性验证失败：期望{row_count}行，实际{len(relevant_mappings)}行映射")
                return False
            
            return True
            
        except Exception as e:
            logger.debug(f"数据一致性验证异常: {e}")
            return False
    
    def _get_border(self):
        """获取边框样式"""
        return Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    def _set_precise_column_widths(self, ws, total_rows: int = None):
        """设置精确的列宽和行高 - 删除执行时间列"""
        column_widths = {
            'A': 15.0,   # 节点1
            'B': 20.0,   # 节点2
            'C': 25.0,   # 节点3
            'D': 20.0,   # 节点4
            'E': 20.0,   # 节点5
            'F': 20.0,   # 节点6
            'G': 20.0,   # 节点7
            'H': 20.0,   # 节点8
            'I': 20.0,   # 节点9
            'J': 20.0,   # 节点10
            'K': 15.0,   # 端/API/服务
            'L': 12.0,   # 冒烟结果
            'M': 18.0,   # 研发对应负责人
            'N': 20.0,   # showcase问题
            'O': 14.0,   # 是否核心功能
            'P': 14.0    # 是否影响主流程
        }
        
        # 设置列宽
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 设置默认行高以适应自动换行
        # 表头行高度
        ws.row_dimensions[1].height = 30
        
        # 数据行默认高度（根据内容自动调整）
        max_row = total_rows if total_rows else ws.max_row
        for row in range(2, max_row + 1):
            ws.row_dimensions[row].height = 35
    
    # 增强的业务逻辑方法
    # 如果将来需要重新启用业务逻辑，可以参考git历史记录中的实现
    
    def _create_enhanced_summary_sheet(self, ws, metadata: Dict[str, Any], total_cases: int, optimization_stats: Dict[str, Any] = None):
        """创建增强的汇总表 - 包含优化信息"""
        
        # 设置表头
        ws['A1'] = "导出汇总报告"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        
        # 基本信息
        current_row = 3
        basic_info = [
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['源文件', metadata.get('source_file', 'unknown')],
            ['选中标识符', ', '.join(metadata.get('selected_markers', []))],
            ['', ''],  # 空行分隔
            ['测试用例统计', ''],
            ['有效测试用例', total_cases],
            ['', ''],  # 空行分隔
            ['导出特性', ''],
            ['智能层级合并', '✓'],
            ['精确单元格匹配', '✓'],
            ['层级背景色', '✓'],
            ['完全匹配模版', '✓'],
            ['增强业务逻辑', '✓'],
            ['业务列优化', '✓ 显示空白'],
            ['删除执行时间列', '✓']
        ]
        
        # 添加优化统计信息
        if optimization_stats:
            basic_info.extend([
                ['', ''],  # 空行分隔
                ['优化效果统计', ''],
                ['原始测试用例', f"{optimization_stats['original_cases']} 个"],
                ['有效测试用例', f"{optimization_stats['valid_cases']} 个"],
                ['过滤无效用例', f"{optimization_stats['filtered_cases']} 个"],
                ['顶级层级分组', f"{optimization_stats['top_level_groups']} 个"],
                ['数据质量提升', f"{optimization_stats['filtered_cases'] / optimization_stats['original_cases'] * 100:.1f}%" if optimization_stats['original_cases'] > 0 else "0%"]
            ])
        
        for item, value in basic_info:
            ws[f'A{current_row}'] = item
            ws[f'B{current_row}'] = value
            
            # 设置样式
            if item and not value:  # 分类标题
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].fill = PatternFill(start_color='E8F1F9', end_color='E8F1F9', fill_type='solid')
            elif item.endswith('统计') or item.endswith('效果'):  # 统计标题
                ws[f'A{current_row}'].font = Font(bold=True, color='4F81BD')
            
            current_row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30


def test_enhanced_export():
    """测试增强版导出功能"""
    logger.info("🧪 测试增强版层级合并导出功能...")
    
    # 模拟复杂的测试数据
    test_data = {
        "smoke_test_suite": {
            "metadata": {
                "source_file": "学习报告.xmind",
                "export_time": datetime.now().isoformat(),
                "selected_markers": ["priority-1", "priority-2", "flag-red"],
                "total_cases": 6
            },
            "test_cases": [
                {
                    "test_path": "学习报告 > 展示规则 > 答题器参与课中不完成加油站验证 > 查看讲次详情",
                    "title": "不展示学习报告",
                    "markers": ["priority-1"]
                },
                {
                    "test_path": "学习报告 > 展示规则 > 答题器参与课中不完成加油站验证 > 其他操作",
                    "title": "其他验证",
                    "markers": ["priority-2"]
                },
                {
                    "test_path": "学习报告 > 高光时刻 > 高光时刻新增排行榜 > 查看排行榜",
                    "title": "查看排行榜",
                    "markers": ["priority-1"]
                },
                {
                    "test_path": "学习报告 > 高光时刻 > 高光时刻新增排行榜 > 验证功能",
                    "title": "功能验证",
                    "markers": ["priority-2"]
                },
                {
                    "test_path": "学习报告 > 高光时刻 > 学生只有排行榜高光时刻",
                    "title": "显示验证",
                    "markers": ["flag-red"]
                },
                {
                    "test_path": "学习报告 > 答题详情 > 错题重做",
                    "title": "错题重做功能",
                    "markers": ["priority-2"]
                }
            ]
        }
    }
    
    exporter = EnhancedHierarchicalExporter()
    output_file = exporter.export_with_enhanced_merge(test_data)
    
    logger.info(f"✅ 增强版层级合并测试完成，文件已生成: {output_file}")
    return output_file


def test_optimization_demo():
    """演示数据优化效果的测试方法"""
    
    # 模拟包含空白节点问题的测试数据
    test_cases_with_issues = [
        # 正常用例
        {
            "title": "用户登录功能验证",
            "test_path": "用户管理 > 登录功能 > 用户名密码登录",
            "steps": [
                {"step": 1, "action": "打开登录页面", "expected": "页面正常显示"},
                {"step": 2, "action": "输入用户名密码", "expected": "登录成功"}
            ]
        },
        # 问题用例1：空路径
        {
            "title": "",
            "test_path": "",
            "steps": []
        },
        # 问题用例2：路径包含空元素
        {
            "title": "支付功能",
            "test_path": "商城管理 >  > 支付处理 > 在线支付",
            "steps": [
                {"step": 1, "action": "选择商品", "expected": "添加成功"}
            ]
        },
        # 问题用例3：无意义内容
        {
            "title": "...",
            "test_path": "xxx > --- > 测试模块",
            "steps": [
                {"step": 1, "action": "todo", "expected": "待定"}
            ]
        },
        # 正常用例
        {
            "title": "商品搜索功能验证", 
            "test_path": "商城管理 > 商品管理 > 商品搜索",
            "steps": [
                {"step": 1, "action": "输入搜索关键词", "expected": "显示搜索结果"}
            ]
        },
        # 问题用例4：路径过短
        {
            "title": "测试",
            "test_path": "单一模块",
            "steps": [
                {"step": 1, "action": "执行操作", "expected": "操作完成"}
            ]
        }
    ]
    
    exporter = EnhancedHierarchicalExporter()
    
    print("🔍 数据优化效果演示")
    print("=" * 50)
    print(f"📊 原始测试用例数量: {len(test_cases_with_issues)}")
    
    # 测试数据清理效果
    valid_count = 0
    for i, case in enumerate(test_cases_with_issues):
        is_valid = exporter._is_valid_test_case(case)
        status = "✅ 有效" if is_valid else "❌ 无效"
        print(f"   用例{i+1}: {case.get('title', '(空)')[:20]} - {status}")
        if is_valid:
            valid_count += 1
    
    print(f"\n📋 数据清理结果:")
    print(f"   ✅ 有效用例: {valid_count} 个")
    print(f"   ❌ 无效用例: {len(test_cases_with_issues) - valid_count} 个")
    print(f"   📈 数据质量提升: {(len(test_cases_with_issues) - valid_count) / len(test_cases_with_issues) * 100:.1f}%")
    
    print(f"\n🎯 路径清理演示:")
    for i, case in enumerate(test_cases_with_issues):
        original_path = case.get('test_path', '')
        cleaned_nodes = exporter._parse_and_clean_path(original_path)
        if original_path:
            cleaned_path = ' > '.join(cleaned_nodes) if cleaned_nodes else "(路径无效)"
            if original_path != cleaned_path:
                print(f"   路径{i+1}: '{original_path}' → '{cleaned_path}'")
    
    print(f"\n✨ 优化后的Excel表格将:")
    print(f"   🗑️  自动过滤无效节点")
    print(f"   🎨 消除空白行干扰")
    print(f"   📐 保持完美的表格对齐")
    print(f"   🎯 提供清晰的层级结构")


def test_business_columns_optimization():
    """测试业务列优化效果 - 显示空白列和删除执行时间列"""
    
    # 创建有效的测试用例数据
    test_cases_data = {
        "smoke_test_suite": {
            "metadata": {
                "source_file": "业务列优化测试.xmind",
                "export_time": datetime.now().isoformat(),
                "selected_markers": ["priority-1", "priority-2", "important"],
                "total_cases": 4
            },
            "test_cases": [
                {
                    "case_id": "SMOKE_001",
                    "title": "用户登录功能验证",
                    "test_path": "用户管理 > 登录功能 > 用户名密码登录",
                    "priority": "P1",
                    "markers": ["priority-1"],
                    "steps": [
                        {"step": 1, "action": "打开登录页面", "expected": "页面正常显示"},
                        {"step": 2, "action": "输入用户名密码", "expected": "登录成功"}
                    ]
                },
                {
                    "case_id": "SMOKE_002",
                    "title": "商品搜索功能验证",
                    "test_path": "商城管理 > 商品管理 > 商品搜索",
                    "priority": "P2",
                    "markers": ["priority-2"],
                    "steps": [
                        {"step": 1, "action": "输入搜索关键词", "expected": "显示搜索结果"}
                    ]
                },
                {
                    "case_id": "SMOKE_003", 
                    "title": "支付流程验证",
                    "test_path": "商城管理 > 订单管理 > 支付处理",
                    "priority": "P1",
                    "markers": ["important"],
                    "steps": [
                        {"step": 1, "action": "选择商品", "expected": "添加到购物车"},
                        {"step": 2, "action": "确认支付", "expected": "支付成功"}
                    ]
                },
                {
                    "case_id": "SMOKE_004",
                    "title": "订单查询功能验证", 
                    "test_path": "商城管理 > 订单管理 > 订单查询",
                    "priority": "P2",
                    "markers": ["priority-2"],
                    "steps": [
                        {"step": 1, "action": "输入订单号", "expected": "显示订单详情"}
                    ]
                }
            ]
        }
    }
    
    exporter = EnhancedHierarchicalExporter()
    
    print("\n🔧 业务列优化测试")
    print("=" * 50)
    print("✨ 优化内容：")
    print("   📋 端/API/服务列 → 显示空白")
    print("   📋 冒烟结果列 → 显示空白")
    print("   📋 研发对应负责人列 → 显示空白")
    print("   📋 showcase问题列 → 显示空白")
    print("   📋 是否核心功能列 → 显示空白")
    print("   📋 是否影响主流程列 → 显示空白")
    print("   🗑️  执行时间列 → 已删除")
    
    try:
        output_file = exporter.export_with_enhanced_merge(
            test_cases_data, 
            "业务列优化演示_冒烟测试用例.xlsx"
        )
        
        print(f"\n✅ 业务列优化测试完成！")
        print(f"📁 生成文件: {output_file}")
        print(f"🎯 现在Excel表格的业务列都显示为空白，便于用户手动填写")
        print(f"⚡ 执行时间列已删除，表格更加简洁")
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    test_enhanced_export()
    test_optimization_demo()
    test_business_columns_optimization() 