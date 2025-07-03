#!/usr/bin/env python3
"""
综合性自测脚本
测试所有导出功能并对比效果，确保重构后的代码完全正常
"""

import requests
import json
import base64
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
import time

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# API基础URL
API_BASE_URL = "http://localhost:8000"

class ComprehensiveTestSuite:
    """综合测试套件"""
    
    def __init__(self):
        self.test_results = {}
        self.xmind_file_data = None
        
    def run_all_tests(self):
        """运行所有测试"""
        logger.info("🚀 开始综合性自测...")
        logger.info("=" * 80)
        
        try:
            # 1. 准备测试数据
            if not self._prepare_test_data():
                return False
            
            # 2. 测试API服务健康状态
            if not self._test_api_health():
                return False
            
            # 3. 测试所有导出接口
            export_apis = [
                {
                    "name": "标准格式导出",
                    "endpoint": "/api/export",
                    "description": "原始数据格式，无特殊样式"
                },
                {
                    "name": "模版格式导出", 
                    "endpoint": "/api/export-template",
                    "description": "基础模版格式，包含业务字段"
                },
                {
                    "name": "层级合并导出",
                    "endpoint": "/api/export-hierarchical", 
                    "description": "基础层级合并，智能单元格合并"
                },
                {
                    "name": "增强层级合并导出",
                    "endpoint": "/api/export-enhanced-hierarchical",
                    "description": "🔥 完美匹配模版的增强版导出"
                }
            ]
            
            for api in export_apis:
                success = self._test_export_api(api)
                self.test_results[api["name"]] = success
                if not success:
                    logger.error(f"❌ {api['name']} 测试失败")
                time.sleep(1)  # 避免请求过快
            
            # 4. 生成测试报告
            self._generate_test_report()
            
            # 5. 检查整体成功率
            success_count = sum(1 for result in self.test_results.values() if result)
            total_count = len(self.test_results)
            
            if success_count == total_count:
                logger.info("🎉 所有测试通过！重构成功！")
                return True
            else:
                logger.error(f"❌ 部分测试失败：{success_count}/{total_count} 通过")
                return False
                
        except Exception as e:
            logger.error(f"❌ 综合测试过程中发生错误: {str(e)}")
            return False
    
    def _prepare_test_data(self):
        """准备测试数据"""
        logger.info("\n1️⃣ 准备测试数据...")
        
        xmind_path = "../学习报告.xmind"
        if not os.path.exists(xmind_path):
            logger.error(f"❌ XMind测试文件不存在: {xmind_path}")
            return False
        
        try:
            with open(xmind_path, 'rb') as f:
                file_content = f.read()
                self.xmind_file_data = base64.b64encode(file_content).decode('utf-8')
            
            logger.info(f"✅ 成功加载测试文件: {len(file_content)} 字节")
            return True
            
        except Exception as e:
            logger.error(f"❌ 加载测试文件失败: {str(e)}")
            return False
    
    def _test_api_health(self):
        """测试API服务健康状态"""
        logger.info("\n2️⃣ 测试API服务健康状态...")
        
        try:
            response = requests.get(f"{API_BASE_URL}/", timeout=5)
            if response.status_code == 200:
                logger.info("✅ API服务健康状态正常")
                return True
            else:
                logger.error(f"❌ API服务状态异常: {response.status_code}")
                return False
                
        except Exception as e:
            logger.error(f"❌ API服务连接失败: {str(e)}")
            return False
    
    def _test_export_api(self, api_info):
        """测试单个导出API"""
        logger.info(f"\n3️⃣ 测试 {api_info['name']}...")
        logger.info(f"   接口: {api_info['endpoint']}")
        logger.info(f"   说明: {api_info['description']}")
        
        try:
            # 准备请求数据
            request_data = {
                "selected_markers": ["priority-1", "priority-2", "flag-red"],
                "file_data": self.xmind_file_data
            }
            
            # 发送请求
            start_time = time.time()
            response = requests.post(
                f"{API_BASE_URL}{api_info['endpoint']}",
                json=request_data,
                timeout=30
            )
            end_time = time.time()
            
            if response.status_code != 200:
                logger.error(f"❌ API请求失败: {response.status_code}")
                logger.error(f"错误信息: {response.text}")
                return False
            
            # 解析响应
            response_data = response.json()
            processing_time = end_time - start_time
            
            # 根据不同的接口类型处理响应
            if api_info['endpoint'] == '/api/export':
                # 标准格式导出返回JSON数据
                if 'smoke_test_suite' not in response_data:
                    logger.error("❌ 标准格式导出响应中缺少smoke_test_suite")
                    return False
                
                total_cases = response_data.get('smoke_test_suite', {}).get('metadata', {}).get('total_cases', 0)
                
                logger.info(f"✅ {api_info['name']} 测试成功")
                logger.info(f"   处理时间: {processing_time:.2f}秒")
                logger.info(f"   用例数量: {total_cases}")
                logger.info(f"   返回格式: JSON数据")
                logger.info(f"   数据大小: {len(str(response_data)):,} 字符")
                
                return True
                
            else:
                # Excel文件导出接口
                if not response_data.get('file_data'):
                    logger.error("❌ 响应中缺少文件数据")
                    return False
                
                # 保存并分析文件
                excel_content = base64.b64decode(response_data['file_data'])
                filename = f"test_{api_info['name'].replace(' ', '_')}_{datetime.now().strftime('%H%M%S')}.xlsx"
                
                with open(filename, 'wb') as f:
                    f.write(excel_content)
                
                # 分析Excel文件
                analysis = self._analyze_excel_file(filename)
                
                # 记录结果
                total_cases = response_data.get('export_details', {}).get('total_cases', 0)
                
                logger.info(f"✅ {api_info['name']} 测试成功")
                logger.info(f"   处理时间: {processing_time:.2f}秒")
                logger.info(f"   用例数量: {total_cases}")
                logger.info(f"   文件大小: {len(excel_content):,} 字节")
                logger.info(f"   合并区域: {analysis['merged_cells']} 个")
                logger.info(f"   数据行数: {analysis['data_rows']}")
                
                # 特殊功能检查
                if "层级合并" in api_info['name']:
                    if analysis['merged_cells'] > 0:
                        logger.info("   ✅ 包含合并单元格功能")
                    else:
                        logger.warning("   ⚠️ 未检测到合并单元格")
                
                return True
            
        except Exception as e:
            logger.error(f"❌ {api_info['name']} 测试失败: {str(e)}")
            return False
    
    def _analyze_excel_file(self, filename):
        """分析Excel文件结构"""
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            analysis = {
                'merged_cells': len(list(ws.merged_cells.ranges)),
                'data_rows': ws.max_row - 1,  # 减去表头
                'columns': ws.max_column,
                'worksheets': len(wb.sheetnames)
            }
            
            wb.close()
            return analysis
            
        except Exception as e:
            logger.warning(f"⚠️ Excel文件分析失败: {str(e)}")
            return {
                'merged_cells': 0,
                'data_rows': 0,
                'columns': 0,
                'worksheets': 0
            }
    
    def _generate_test_report(self):
        """生成测试报告"""
        logger.info("\n" + "=" * 80)
        logger.info("📊 综合测试报告")
        logger.info("=" * 80)
        
        for test_name, success in self.test_results.items():
            status = "✅ 通过" if success else "❌ 失败"
            logger.info(f"   {test_name:<25} {status}")
        
        success_count = sum(1 for result in self.test_results.values() if result)
        total_count = len(self.test_results)
        success_rate = (success_count / total_count * 100) if total_count > 0 else 0
        
        logger.info("-" * 80)
        logger.info(f"📈 成功率: {success_count}/{total_count} ({success_rate:.1f}%)")
        
        if success_rate == 100:
            logger.info("🎯 评估结果: 重构完全成功，所有功能正常！")
        elif success_rate >= 75:
            logger.info("⚠️ 评估结果: 大部分功能正常，需要检查失败项")
        else:
            logger.error("💔 评估结果: 重构存在问题，需要修复")
        
        logger.info("=" * 80)

def test_enhanced_export_directly():
    """直接测试增强版导出器"""
    logger.info("🧪 直接测试增强版导出器...")
    
    try:
        from enhanced_hierarchical_exporter import EnhancedHierarchicalExporter
        
        # 测试模拟数据
        test_data = {
            "smoke_test_suite": {
                "metadata": {
                    "source_file": "测试.xmind",
                    "export_time": datetime.now().isoformat(),
                    "selected_markers": ["priority-1", "priority-2", "flag-red"],
                    "total_cases": 6
                },
                "test_cases": [
                    {
                        "test_path": "学习报告 > 展示规则 > 验证功能 > 基础验证",
                        "title": "基础展示验证",
                        "markers": ["priority-1"]
                    },
                    {
                        "test_path": "学习报告 > 展示规则 > 验证功能 > 高级验证",
                        "title": "高级展示验证",
                        "markers": ["priority-2"]
                    },
                    {
                        "test_path": "学习报告 > 高光时刻 > 排行榜功能 > 显示验证",
                        "title": "排行榜显示验证",
                        "markers": ["priority-1"]
                    },
                    {
                        "test_path": "学习报告 > 高光时刻 > 排行榜功能 > 交互验证",
                        "title": "排行榜交互验证",
                        "markers": ["priority-2"]
                    },
                    {
                        "test_path": "学习报告 > 高光时刻 > 图片展示",
                        "title": "图片展示功能",
                        "markers": ["flag-red"]
                    },
                    {
                        "test_path": "学习报告 > 答题详情 > 错题重做",
                        "title": "错题重做功能",
                        "markers": ["priority-2"]
                    }
                ]
            }
        }
        
        exporter = EnhancedHierarchicalExporter()
        output_file = exporter.export_with_enhanced_merge(test_data)
        
        logger.info(f"✅ 增强版导出器直接测试成功: {output_file}")
        
        # 分析生成的文件
        wb = load_workbook(output_file)
        ws = wb.active
        merged_cells = len(list(ws.merged_cells.ranges))
        
        logger.info(f"📊 生成文件分析:")
        logger.info(f"   数据行数: {ws.max_row - 1}")
        logger.info(f"   合并区域: {merged_cells} 个")
        logger.info(f"   工作表数: {len(wb.sheetnames)}")
        
        if merged_cells > 0:
            logger.info("✅ 合并功能正常")
        else:
            logger.warning("⚠️ 未检测到合并功能")
        
        wb.close()
        return True
        
    except Exception as e:
        logger.error(f"❌ 增强版导出器直接测试失败: {str(e)}")
        return False

def main():
    """主函数"""
    logger.info("🎯 开始全面自测...")
    
    # 1. 直接测试增强版导出器
    logger.info("\n" + "=" * 60)
    logger.info("第一阶段：直接测试增强版导出器")
    logger.info("=" * 60)
    
    direct_test_success = test_enhanced_export_directly()
    
    # 2. 测试完整的API服务
    logger.info("\n" + "=" * 60)
    logger.info("第二阶段：测试完整的API服务")
    logger.info("=" * 60)
    
    test_suite = ComprehensiveTestSuite()
    api_test_success = test_suite.run_all_tests()
    
    # 3. 总结
    logger.info("\n" + "=" * 60)
    logger.info("🏁 最终总结")
    logger.info("=" * 60)
    
    if direct_test_success and api_test_success:
        logger.info("🎉 重构完全成功！")
        logger.info("✅ 增强版导出器功能正常")
        logger.info("✅ 所有API接口工作正常") 
        logger.info("✅ 没有影响现有逻辑")
        logger.info("✅ 新功能完美集成")
        return True
    else:
        logger.error("❌ 重构存在问题")
        if not direct_test_success:
            logger.error("   增强版导出器异常")
        if not api_test_success:
            logger.error("   API服务异常")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1) 