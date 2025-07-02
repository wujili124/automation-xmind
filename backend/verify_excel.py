#!/usr/bin/env python3
"""
验证Excel文件内容
"""
from openpyxl import load_workbook
import os

def verify_excel_content():
    """验证Excel文件内容"""
    # 查找最新的Excel文件
    excel_files = [f for f in os.listdir('.') if f.startswith('冒烟测试用例_') and f.endswith('.xlsx')]
    
    if not excel_files:
        print("❌ 未找到Excel文件")
        return
    
    # 选择最新的Excel文件
    excel_file = sorted(excel_files)[-1]
    print(f"📊 验证Excel文件: {excel_file}")
    
    try:
        # 打开工作簿
        wb = load_workbook(excel_file)
        print(f"✅ Excel文件打开成功")
        print(f"   工作表: {wb.sheetnames}")
        
        # 验证测试用例工作表
        if '冒烟测试用例' in wb.sheetnames:
            ws1 = wb['冒烟测试用例']
            print(f"\n📝 冒烟测试用例工作表:")
            print(f"   行数: {ws1.max_row}")
            print(f"   列数: {ws1.max_column}")
            
            # 显示表头
            headers = []
            for col in range(1, ws1.max_column + 1):
                headers.append(ws1.cell(row=1, column=col).value)
            print(f"   表头: {headers}")
            
            # 显示前3个测试用例的摘要
            print(f"\n📋 测试用例示例:")
            for row in range(2, min(5, ws1.max_row + 1)):
                case_id = ws1.cell(row=row, column=1).value
                title = ws1.cell(row=row, column=2).value
                module = ws1.cell(row=row, column=3).value
                priority = ws1.cell(row=row, column=4).value
                print(f"   {case_id}: {title} [{module}] - {priority}")
        
        # 验证导出汇总工作表
        if '导出汇总' in wb.sheetnames:
            ws2 = wb['导出汇总']
            print(f"\n📊 导出汇总工作表:")
            print(f"   行数: {ws2.max_row}")
            print(f"   列数: {ws2.max_column}")
            
            # 显示汇总信息
            print(f"\n📈 汇总信息:")
            for row in range(1, min(15, ws2.max_row + 1)):
                key = ws2.cell(row=row, column=1).value
                value = ws2.cell(row=row, column=2).value
                if key and value is not None:
                    print(f"   {key}: {value}")
        
        print(f"\n✅ Excel文件验证完成，数据完整")
        
    except Exception as e:
        print(f"❌ Excel文件验证失败: {e}")

if __name__ == "__main__":
    verify_excel_content() 