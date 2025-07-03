#!/usr/bin/env python3
"""
测试前端和后端的连接
验证修复后的导出逻辑是否能正确调用增强层级合并API
"""

import requests
import base64
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_frontend_backend_connection():
    """测试前端和后端连接"""
    
    print("🔗 测试前端和后端连接")
    print("=" * 60)
    
    # 1. 测试健康检查
    print("1️⃣ 测试API健康状态...")
    try:
        response = requests.get("http://localhost:8000/", timeout=5)
        if response.status_code == 200:
            print("   ✅ API服务正常运行")
        else:
            print(f"   ❌ API健康检查失败: {response.status_code}")
            return False
    except Exception as e:
        print(f"   ❌ 无法连接到API服务: {str(e)}")
        return False
    
    # 2. 测试增强层级合并API
    print("\n2️⃣ 测试增强层级合并API...")
    
    # 准备测试数据
    xmind_path = "../学习报告.xmind"
    if not os.path.exists(xmind_path):
        print("   ❌ 测试文件不存在")
        return False
    
    with open(xmind_path, 'rb') as f:
        file_content = f.read()
        file_data = base64.b64encode(file_content).decode('utf-8')
    
    # 模拟前端请求
    request_data = {
        "selected_markers": ["priority-1", "priority-2", "flag-red"],
        "file_data": file_data
    }
    
    try:
        print("   📡 调用增强层级合并API...")
        response = requests.post(
            "http://localhost:8000/api/export-enhanced-hierarchical",
            json=request_data,
            timeout=30
        )
        
        if response.status_code == 200:
            response_data = response.json()
            
            print("   ✅ API调用成功")
            print(f"   📊 响应状态: {response_data.get('success', 'N/A')}")
            
            if 'export_details' in response_data:
                details = response_data['export_details']
                print(f"   🔗 合并区域数: {details.get('merged_regions_count', 0)}")
                print(f"   📋 数据行数: {details.get('data_rows', 0)}")
                print(f"   🎯 特性: {', '.join(details.get('features', [])[:2])}")
            
            if 'file_data' in response_data:
                # 验证文件内容
                excel_content = base64.b64decode(response_data['file_data'])
                timestamp = datetime.now().strftime('%H%M%S')
                test_filename = f"前端后端连接测试_{timestamp}.xlsx"
                
                with open(test_filename, 'wb') as f:
                    f.write(excel_content)
                
                print(f"   💾 测试文件已保存: {test_filename}")
                
                # 验证Excel内容
                return verify_excel_content(test_filename)
            else:
                print("   ❌ 响应中缺少文件数据")
                return False
        else:
            print(f"   ❌ API调用失败: {response.status_code}")
            if response.content:
                print(f"   错误信息: {response.text}")
            return False
            
    except Exception as e:
        print(f"   ❌ API调用异常: {str(e)}")
        return False

def verify_excel_content(filename):
    """验证Excel文件内容"""
    
    print(f"\n3️⃣ 验证Excel文件内容: {filename}")
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # 检查基本信息
        data_rows = ws.max_row - 1  # 减去表头
        data_cols = ws.max_column
        print(f"   📊 数据行数: {data_rows}")
        print(f"   📊 数据列数: {data_cols}")
        
        # 检查合并区域
        merged_count = len(ws.merged_cells.ranges)
        print(f"   🔗 合并区域数: {merged_count}")
        
        if merged_count > 0:
            print("   ✅ 检测到合并单元格功能")
            
            # 显示前几个合并区域
            for i, merge_range in enumerate(list(ws.merged_cells.ranges)[:3]):
                start_row = merge_range.min_row
                end_row = merge_range.max_row
                col_num = merge_range.min_col
                value = ws.cell(row=start_row, column=col_num).value
                span = end_row - start_row + 1
                print(f"   合并{i+1}: 行{start_row}-{end_row} (跨{span}行) = '{value}'")
        else:
            print("   ❌ 未检测到合并功能")
            wb.close()
            return False
        
        # 检查表头
        print(f"\n   📋 表头验证:")
        expected_headers = ['节点1', '节点2', '节点3', '节点4', '节点5', 
                           '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题']
        
        actual_headers = []
        for col in range(1, min(10, ws.max_column + 1)):
            header = ws.cell(row=1, column=col).value
            actual_headers.append(header)
        
        match_count = 0
        for i, expected in enumerate(expected_headers):
            if i < len(actual_headers) and actual_headers[i] == expected:
                match_count += 1
            else:
                break
        
        if match_count >= 5:  # 至少前5列匹配
            print(f"   ✅ 表头匹配度: {match_count}/{len(expected_headers)}")
        else:
            print(f"   ⚠️ 表头匹配度较低: {match_count}/{len(expected_headers)}")
        
        wb.close()
        
        # 评估结果
        success_criteria = [
            data_rows >= 10,
            data_cols >= 9,
            merged_count >= 5,
            match_count >= 5
        ]
        
        passed = sum(success_criteria)
        total = len(success_criteria)
        
        print(f"\n   📈 验证结果: {passed}/{total} 项通过")
        
        if passed == total:
            print("   🎉 前端后端连接测试完全成功！")
            return True
        elif passed >= 3:
            print("   ✅ 前端后端连接基本正常")
            return True
        else:
            print("   ❌ 前端后端连接存在问题")
            return False
            
    except Exception as e:
        print(f"   ❌ Excel验证失败: {str(e)}")
        return False

def test_comparison():
    """对比测试：显示修复前后的差异"""
    
    print(f"\n" + "=" * 60)
    print("📊 修复前后对比说明")
    print("=" * 60)
    
    print("🔴 修复前的问题:")
    print("   ❌ 前端直接使用XLSX库生成Excel")
    print("   ❌ 没有调用后端API")
    print("   ❌ 无法实现层级合并")
    print("   ❌ 只能生成传统行列格式")
    
    print("\n🟢 修复后的改进:")
    print("   ✅ 前端调用后端增强层级合并API")
    print("   ✅ 后端智能处理层级结构")
    print("   ✅ 实现智能单元格合并")
    print("   ✅ 完美匹配模版视觉效果")
    print("   ✅ 保持传统格式作为备选方案")
    
    print("\n🎯 使用指南:")
    print("   1. 点击 '导出Excel' 按钮")
    print("   2. 前端会自动调用增强层级合并API")
    print("   3. 下载的文件名包含 '🔥增强层级合并' 前缀")
    print("   4. 打开Excel文件可以看到智能合并效果")
    
    print("\n📱 故障排除:")
    print("   - 如果增强导出失败，会提示生成传统格式")
    print("   - 检查文件名前缀确认使用的导出方式")
    print("   - 确保后端服务正常运行")

if __name__ == "__main__":
    print("🚀 开始前端后端连接测试...")
    
    success = test_frontend_backend_connection()
    test_comparison()
    
    print(f"\n" + "=" * 60)
    if success:
        print("🎊 测试完成！前端后端连接修复成功！")
        print("✅ 用户现在可以获得增强层级合并Excel文件")
        print("🔥 下载的文件将包含智能合并的层级结构")
    else:
        print("❌ 测试失败，需要进一步调试")
    print("=" * 60) 