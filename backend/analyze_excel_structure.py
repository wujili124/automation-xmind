#!/usr/bin/env python3
"""
详细分析《冒烟用例导出模版.xlsx》的结构
重点分析表格合并、节点排列和视觉效果
"""

import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def analyze_excel_structure():
    """详细分析Excel模版的结构和合并单元格"""
    
    template_path = "../冒烟用例导出模版.xlsx"
    
    if not os.path.exists(template_path):
        logger.error(f"❌ 模版文件不存在: {template_path}")
        return None
    
    logger.info("🔍 开始详细分析Excel模版结构...")
    
    try:
        # 使用openpyxl详细分析
        wb = load_workbook(template_path)
        
        for sheet_name in wb.sheetnames:
            logger.info(f"\n📋 分析工作表: {sheet_name}")
            ws = wb[sheet_name]
            
            if sheet_name == "冒烟测试用例":
                analyze_main_sheet_structure(ws)
            elif sheet_name == "导出汇总":
                analyze_summary_sheet_structure(ws)
        
        wb.close()
        
    except Exception as e:
        logger.error(f"❌ 分析失败: {str(e)}")
        return None

def analyze_main_sheet_structure(ws):
    """详细分析主工作表的结构"""
    logger.info("🎯 分析冒烟测试用例工作表结构...")
    
    # 1. 分析合并单元格
    logger.info("\n🔗 合并单元格分析:")
    merged_ranges = list(ws.merged_cells.ranges)
    
    if merged_ranges:
        for i, merged_range in enumerate(merged_ranges):
            start_row, start_col = merged_range.min_row, merged_range.min_col
            end_row, end_col = merged_range.max_row, merged_range.max_col
            
            # 获取合并单元格的值
            cell_value = ws.cell(start_row, start_col).value
            
            logger.info(f"   合并{i+1}: {merged_range} = {cell_value}")
            logger.info(f"     范围: 行{start_row}-{end_row}, 列{start_col}-{end_col}")
            logger.info(f"     尺寸: {end_row-start_row+1}行 x {end_col-start_col+1}列")
            
            # 分析合并模式
            if end_row - start_row > 0:
                logger.info(f"     ✅ 垂直合并: 跨{end_row-start_row+1}行")
            if end_col - start_col > 0:
                logger.info(f"     ✅ 水平合并: 跨{end_col-start_col+1}列")
    else:
        logger.info("   ❌ 没有发现合并单元格")
    
    # 2. 分析表头结构
    logger.info("\n📊 表头结构分析:")
    for row in range(1, min(ws.max_row + 1, 6)):  # 分析前5行
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            row_data.append(cell_value)
        
        logger.info(f"   第{row}行: {row_data}")
    
    # 3. 分析数据排列模式
    logger.info("\n📈 数据排列模式分析:")
    
    # 查找节点列的模式
    node_columns = {}
    for col in range(1, 6):  # 前5列通常是节点
        col_letter = get_column_letter(col)
        col_values = []
        
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            cell_value = ws.cell(row, col).value
            if cell_value:
                col_values.append(cell_value)
        
        if col_values:
            node_columns[f"节点{col}"] = {
                'unique_values': list(set(col_values)),
                'total_values': len(col_values),
                'pattern_analysis': analyze_column_pattern(col_values)
            }
    
    # 显示节点分析结果
    for node_name, info in node_columns.items():
        logger.info(f"   {node_name}:")
        logger.info(f"     唯一值: {info['unique_values']}")
        logger.info(f"     总数据: {info['total_values']}")
        logger.info(f"     模式: {info['pattern_analysis']}")
    
    # 4. 分析层级关系
    logger.info("\n🏗️ 层级关系分析:")
    analyze_hierarchical_structure(ws)

def analyze_column_pattern(values):
    """分析列的数据模式"""
    if not values:
        return "空列"
    
    # 检查重复模式
    unique_count = len(set(values))
    total_count = len(values)
    
    if unique_count == 1:
        return f"单一值重复 ({values[0]})"
    elif unique_count == total_count:
        return "每行唯一值"
    else:
        return f"部分重复 ({unique_count}/{total_count})"

def analyze_hierarchical_structure(ws):
    """分析层级结构和合并逻辑"""
    
    # 构建数据结构
    data_structure = []
    
    for row in range(2, ws.max_row + 1):  # 从数据行开始
        row_data = {}
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell_value = ws.cell(row, col).value
            
            # 获取表头名称
            header = ws.cell(1, col).value or f"列{col}"
            row_data[header] = cell_value
        
        data_structure.append(row_data)
    
    # 分析节点层级关系
    logger.info("   节点层级分析:")
    
    node_hierarchy = {}
    for row_data in data_structure:
        nodes = []
        for i in range(1, 6):  # 节点1-5
            node_key = f"节点{i}"
            if node_key in row_data and row_data[node_key]:
                nodes.append(row_data[node_key])
        
        if nodes:
            path = " > ".join(nodes)
            if path not in node_hierarchy:
                node_hierarchy[path] = []
            node_hierarchy[path].append(row_data)
    
    # 显示层级关系
    for path, rows in node_hierarchy.items():
        logger.info(f"     路径: {path}")
        logger.info(f"       数据行数: {len(rows)}")
        
        # 分析这个路径下的变化
        varying_fields = analyze_varying_fields(rows)
        if varying_fields:
            logger.info(f"       变化字段: {varying_fields}")

def analyze_varying_fields(rows):
    """分析在相同路径下哪些字段有变化"""
    if len(rows) <= 1:
        return []
    
    varying_fields = []
    all_keys = rows[0].keys()
    
    for key in all_keys:
        values = [row.get(key) for row in rows]
        unique_values = set(v for v in values if v is not None)
        
        if len(unique_values) > 1:
            varying_fields.append(f"{key}({len(unique_values)}种)")
    
    return varying_fields

def analyze_summary_sheet_structure(ws):
    """分析汇总工作表结构"""
    logger.info("📋 分析导出汇总工作表结构...")
    
    # 简单分析汇总表
    for row in range(1, ws.max_row + 1):
        col1_value = ws.cell(row, 1).value
        col2_value = ws.cell(row, 2).value
        logger.info(f"   第{row}行: {col1_value} | {col2_value}")

def generate_structure_recommendations():
    """生成结构改进建议"""
    logger.info("\n💡 结构改进建议:")
    
    recommendations = [
        "1. 表格合并策略:",
        "   - 相同节点1的行应该垂直合并",
        "   - 相同节点1+节点2的组合应该合并",
        "   - 以此类推到节点5",
        "",
        "2. 视觉层级效果:",
        "   - 使用不同的背景色区分层级",
        "   - 节点1使用深色背景",
        "   - 节点2使用中等背景",
        "   - 节点3-5使用浅色背景",
        "",
        "3. 合并单元格实现:",
        "   - 使用openpyxl的merge_cells功能",
        "   - 按层级递归合并相同值的单元格",
        "   - 保持数据完整性",
        "",
        "4. 排序和分组:",
        "   - 按节点1分组",
        "   - 节点1内按节点2分组",
        "   - 以此类推实现层级排序"
    ]
    
    for recommendation in recommendations:
        logger.info(recommendation)

def main():
    """主函数"""
    logger.info("🚀 开始Excel模版结构详细分析...")
    
    analyze_excel_structure()
    generate_structure_recommendations()
    
    logger.info("\n✅ 分析完成！")

if __name__ == "__main__":
    main() 