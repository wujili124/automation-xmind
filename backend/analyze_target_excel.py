#!/usr/bin/env python3
"""
分析目标Excel文件的详细结构
确保导出格式严格匹配
"""

import pandas as pd
import logging
from openpyxl import load_workbook
import os

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def analyze_target_excel(file_path):
    """详细分析目标Excel文件"""
    logger.info(f"🔍 分析目标Excel文件: {file_path}")
    
    if not os.path.exists(file_path):
        logger.error(f"❌ 文件不存在: {file_path}")
        return None
    
    try:
        # 使用openpyxl分析
        wb = load_workbook(file_path)
        logger.info(f"📁 工作表列表: {wb.sheetnames}")
        
        result = {}
        
        # 分析每个工作表
        for sheet_name in wb.sheetnames:
            logger.info(f"\n📊 分析工作表: {sheet_name}")
            ws = wb[sheet_name]
            
            sheet_info = {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'headers': [],
                'sample_data': [],
                'column_widths': {}
            }
            
            # 获取表头
            if ws.max_row > 0:
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(cell_value)
                    
                    # 获取列宽
                    col_letter = ws.cell(row=1, column=col).column_letter
                    col_width = ws.column_dimensions[col_letter].width
                    sheet_info['column_widths'][col_letter] = col_width
                
                sheet_info['headers'] = headers
                logger.info(f"   表头 ({len(headers)}列): {headers}")
                
                # 获取前5行数据作为样例
                sample_rows = min(6, ws.max_row)  # 表头 + 5行数据
                for row in range(1, sample_rows + 1):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(cell_value)
                    sheet_info['sample_data'].append(row_data)
                
                logger.info(f"   数据行数: {ws.max_row - 1}")
                logger.info(f"   列宽信息: {sheet_info['column_widths']}")
                
                # 显示前3行数据
                for i, row_data in enumerate(sheet_info['sample_data'][:3]):
                    logger.info(f"   第{i+1}行: {row_data}")
            
            result[sheet_name] = sheet_info
        
        wb.close()
        
        # 使用pandas再次验证
        logger.info(f"\n📈 使用pandas验证数据:")
        xls = pd.ExcelFile(file_path)
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logger.info(f"   {sheet_name}: {len(df)} 行 x {len(df.columns)} 列")
            
            # 显示数据类型
            logger.info(f"   列信息:")
            for col in df.columns:
                non_null = df[col].count()
                total = len(df)
                dtype = df[col].dtype
                logger.info(f"     {col}: {non_null}/{total} 非空, 类型: {dtype}")
        
        return result
        
    except Exception as e:
        logger.error(f"❌ 分析Excel文件失败: {str(e)}")
        return None

def compare_with_template():
    """对比原始模版和生成的文件"""
    logger.info("\n🔄 对比分析...")
    
    # 分析原始模版
    template_path = "../冒烟用例导出模版.xlsx"
    logger.info("=" * 60)
    logger.info("📋 原始模版文件分析:")
    template_result = analyze_target_excel(template_path)
    
    # 分析生成的文件
    target_path = "冒烟测试用例_模版格式_2025-07-02_23-19-01.xlsx"
    logger.info("\n" + "=" * 60)
    logger.info("🎯 生成的目标文件分析:")
    target_result = analyze_target_excel(target_path)
    
    # 对比分析
    if template_result and target_result:
        logger.info("\n" + "=" * 60)
        logger.info("🔍 对比结果:")
        
        for sheet_name in template_result.keys():
            if sheet_name in target_result:
                template_sheet = template_result[sheet_name]
                target_sheet = target_result[sheet_name]
                
                logger.info(f"\n📊 工作表: {sheet_name}")
                
                # 对比表头
                template_headers = template_sheet['headers']
                target_headers = target_sheet['headers']
                
                if template_headers == target_headers:
                    logger.info("   ✅ 表头完全匹配")
                else:
                    logger.info("   ⚠️ 表头不匹配:")
                    logger.info(f"     模版: {template_headers}")
                    logger.info(f"     生成: {target_headers}")
                
                # 对比列数
                if len(template_headers) == len(target_headers):
                    logger.info(f"   ✅ 列数匹配: {len(template_headers)}")
                else:
                    logger.info(f"   ⚠️ 列数不匹配: 模版{len(template_headers)} vs 生成{len(target_headers)}")
            else:
                logger.info(f"   ❌ 生成文件缺少工作表: {sheet_name}")
    
    return template_result, target_result

def main():
    """主函数"""
    logger.info("🚀 开始分析目标Excel文件...")
    
    # 对比分析
    template_result, target_result = compare_with_template()
    
    # 生成详细的规范
    logger.info("\n" + "=" * 60)
    logger.info("📐 严格的导出规范:")
    
    if target_result:
        for sheet_name, sheet_info in target_result.items():
            logger.info(f"\n📋 工作表: {sheet_name}")
            logger.info(f"   表头: {sheet_info['headers']}")
            logger.info(f"   列数: {sheet_info['max_column']}")
            logger.info(f"   数据行数: {sheet_info['max_row'] - 1}")
            
            if sheet_info['column_widths']:
                logger.info(f"   列宽设置:")
                for col_letter, width in sheet_info['column_widths'].items():
                    if width and width != 8.43:  # 默认列宽
                        logger.info(f"     列 {col_letter}: {width}")
    
    logger.info("\n✅ 分析完成！")

if __name__ == "__main__":
    main() 