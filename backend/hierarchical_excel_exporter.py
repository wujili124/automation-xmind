#!/usr/bin/env python3
"""
层级化Excel导出器
实现智能的单元格合并，完全匹配《冒烟用例导出模版.xlsx》的视觉效果
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Tuple
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

logger = logging.getLogger(__name__)

class HierarchicalExcelExporter:
    """支持层级合并的Excel导出器"""
    
    def __init__(self):
        # 层级背景色配置
        self.level_colors = {
            1: 'D5E8F5',  # 节点1 - 深蓝色
            2: 'E8F1F9',  # 节点2 - 中蓝色  
            3: 'F0F8FF',  # 节点3 - 浅蓝色
            4: 'F8FBFF',  # 节点4 - 极浅蓝
            5: 'FFFFFF',  # 节点5 - 白色
            6: 'FFFFFF',  # 节点6 - 白色
            7: 'FFFFFF',  # 节点7 - 白色
            8: 'FFFFFF',  # 节点8 - 白色
            9: 'FFFFFF',  # 节点9 - 白色
            10: 'FFFFFF'  # 节点10 - 白色
        }
        
        # 其他列的背景色
        self.other_column_color = 'FAFAFA'  # 浅灰色
    
    def export_with_hierarchical_merge(self, test_cases_data: Dict[str, Any], output_path: str = None) -> str:
        """
        按照层级合并导出Excel
        
        Args:
            test_cases_data: 测试用例数据
            output_path: 输出文件路径
            
        Returns:
            生成的文件路径
        """
        try:
            logger.info("🚀 开始按层级合并导出Excel...")
            
            test_cases = test_cases_data['smoke_test_suite']['test_cases']
            metadata = test_cases_data['smoke_test_suite']['metadata']
            
            # 1. 数据预处理和分组
            grouped_data = self._group_data_hierarchically(test_cases)
            logger.info(f"数据分组完成，共 {len(grouped_data)} 个顶级分组")
            
            # 2. 创建工作簿
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = "冒烟测试用例"
            
            # 3. 写入表头
            self._write_headers(ws_main)
            
            # 4. 写入分组数据并合并单元格
            current_row = 2  # 从第2行开始写数据
            total_rows = self._write_hierarchical_data(ws_main, grouped_data, current_row)
            
            # 5. 设置列宽
            self._set_column_widths(ws_main)
            
            # 6. 创建汇总表
            ws_summary = wb.create_sheet("导出汇总")
            self._create_summary_sheet(ws_summary, metadata, len(test_cases))
            
            # 7. 保存文件
            if not output_path:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                output_path = f"层级合并_冒烟测试用例_{timestamp}.xlsx"
            
            wb.save(output_path)
            
            logger.info(f"✅ 层级合并Excel导出完成: {output_path}")
            logger.info(f"   📊 包含 {len(test_cases)} 个测试用例")
            logger.info(f"   🎯 实现了完整的层级合并效果")
            
            return output_path
            
        except Exception as e:
            logger.error(f"❌ 层级合并导出失败: {str(e)}")
            raise Exception(f"层级合并导出失败: {str(e)}")
    
    def _group_data_hierarchically(self, test_cases: List[Dict]) -> OrderedDict:
        """按层级分组数据"""
        
        # 构建层级字典
        hierarchy = OrderedDict()
        
        for case in test_cases:
            # 解析路径
            path_str = case.get('test_path', '') or case.get('测试路径', '') or case.get('title', '')
            
            if ' > ' in path_str:
                nodes = path_str.split(' > ')
            else:
                # 如果没有路径分隔符，使用标题作为最后一个节点
                nodes = [path_str] if path_str else ['未分类']
            
            # 确保至少有一个节点
            if not nodes or not nodes[0]:
                nodes = ['未分类']
            
            # 构建嵌套字典
            current_level = hierarchy
            full_path = []
            
            for i, node in enumerate(nodes):
                full_path.append(node)
                
                if node not in current_level:
                    current_level[node] = {
                        '_data': [],  # 存储数据
                        '_children': OrderedDict(),  # 子节点
                        '_level': i + 1,  # 层级
                        '_full_path': ' > '.join(full_path)  # 完整路径
                    }
                
                # 如果是最后一个节点，添加数据
                if i == len(nodes) - 1:
                    current_level[node]['_data'].append(case)
                
                current_level = current_level[node]['_children']
        
        return hierarchy
    
    def _write_headers(self, ws):
        """写入表头"""
        headers = [
            '节点1', '节点2', '节点3', '节点4', '节点5', '节点6', '节点7', '节点8', '节点9', '节点10',
            '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题',
            '是否核心功能', '是否影响主流程', '执行时间'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            # 白色字体
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            
            # 添加边框
            cell.border = self._get_border()
    
    def _write_hierarchical_data(self, ws, hierarchy: OrderedDict, start_row: int) -> int:
        """写入层级数据并实现智能合并"""
        
        current_row = start_row
        
        # 第一次遍历：收集所有数据行和位置信息
        all_rows = []
        self._collect_all_rows(hierarchy, all_rows, [], 1)
        
        # 第二次遍历：写入数据
        for row_info in all_rows:
            self._write_single_row(ws, current_row, row_info)
            current_row += 1
        
        # 第三次遍历：计算并应用合并
        self._apply_merges(ws, all_rows, start_row)
        
        return current_row
    
    def _collect_all_rows(self, node_dict: OrderedDict, all_rows: List, path_values: List, level: int):
        """递归收集所有数据行"""
        
        for node_name, node_info in node_dict.items():
            current_path = path_values + [node_name]
            
            # 如果有数据，添加数据行
            if node_info['_data']:
                for case_data in node_info['_data']:
                    row_info = {
                        'path_values': current_path.copy(),
                        'level': level,
                        'data': case_data,
                        'full_path': node_info['_full_path']
                    }
                    all_rows.append(row_info)
            
            # 递归处理子节点
            if node_info['_children']:
                self._collect_all_rows(node_info['_children'], all_rows, current_path, level + 1)
    
    def _write_single_row(self, ws, row: int, row_info: Dict):
        """写入单行数据"""
        
        # 写入节点列 (1-10列)
        for col in range(1, 11):
            if col <= len(row_info['path_values']):
                value = row_info['path_values'][col-1]
                level = col
            else:
                value = ''
                level = 10
            
            cell = ws.cell(row=row, column=col, value=value)
            
            # 设置层级背景色
            if value:
                color = self.level_colors.get(level, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self._get_border()
        
        # 写入业务列 (11-17列)
        case_data = row_info['data']
        business_data = [
            self._determine_platform(case_data),        # 端/API/服务
            self._determine_smoke_result(case_data),     # 冒烟结果
            self._determine_developer(case_data),        # 研发对应负责人
            self._determine_showcase_issue(case_data),   # showcase问题
            '是' if case_data.get('是否核心功能') == '是' else '否',  # 是否核心功能
            '是' if case_data.get('是否影响主流程') == '是' else '否', # 是否影响主流程
            case_data.get('执行时间', '< 2分钟')          # 执行时间
        ]
        
        for col, value in enumerate(business_data, 11):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = PatternFill(start_color=self.other_column_color, end_color=self.other_column_color, fill_type='solid')
            cell.border = self._get_border()
            
            # 冒烟结果特殊颜色
            if col == 12:  # 冒烟结果列
                if value == '通过':
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                elif value == '需关注':
                    cell.fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
                elif value == '失败':
                    cell.fill = PatternFill(start_color='FFE8E8', end_color='FFE8E8', fill_type='solid')
    
    def _apply_merges(self, ws, all_rows: List, start_row: int):
        """应用智能合并"""
        
        # 按列分别处理合并
        for col in range(1, 11):  # 对所有节点列进行合并
            self._merge_column(ws, all_rows, col, start_row)
    
    def _merge_column(self, ws, all_rows: List, col: int, start_row: int):
        """合并指定列的相同值"""
        
        if not all_rows:
            return
        
        current_value = None
        merge_start = start_row
        
        for i, row_info in enumerate(all_rows):
            row_num = start_row + i
            
            # 获取当前行在这一列的值
            if col <= len(row_info['path_values']):
                value = row_info['path_values'][col-1]
                
                # 还需要检查前面的路径是否一致，确保只在同一分支内合并
                path_prefix = row_info['path_values'][:col-1]
            else:
                value = ''
                path_prefix = row_info['path_values']
            
            # 检查是否需要开始新的合并区域
            if current_value != value or (i > 0 and self._path_changed(all_rows[i-1], row_info, col-1)):
                # 应用之前的合并
                if current_value and merge_start < row_num - 1:
                    try:
                        ws.merge_cells(f'{get_column_letter(col)}{merge_start}:{get_column_letter(col)}{row_num-1}')
                        logger.debug(f"合并 {get_column_letter(col)}{merge_start}:{get_column_letter(col)}{row_num-1} = {current_value}")
                    except Exception as e:
                        logger.warning(f"合并失败: {e}")
                
                # 开始新的合并区域
                current_value = value
                merge_start = row_num
        
        # 处理最后一个合并区域
        if current_value and merge_start < start_row + len(all_rows) - 1:
            try:
                end_row = start_row + len(all_rows) - 1
                ws.merge_cells(f'{get_column_letter(col)}{merge_start}:{get_column_letter(col)}{end_row}')
                logger.debug(f"最后合并 {get_column_letter(col)}{merge_start}:{get_column_letter(col)}{end_row} = {current_value}")
            except Exception as e:
                logger.warning(f"最后合并失败: {e}")
    
    def _path_changed(self, prev_row: Dict, curr_row: Dict, level: int) -> bool:
        """检查指定层级的路径前缀是否发生变化"""
        if level <= 0:
            return False
        
        prev_prefix = prev_row['path_values'][:level]
        curr_prefix = curr_row['path_values'][:level]
        
        return prev_prefix != curr_prefix
    
    def _get_border(self):
        """获取边框样式"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _set_column_widths(self, ws):
        """设置列宽"""
        column_widths = {
            'A': 15.0,  # 节点1
            'B': 20.0,  # 节点2
            'C': 25.0,  # 节点3
            'D': 20.0,  # 节点4
            'E': 20.0,  # 节点5
            'F': 20.0,  # 节点6
            'G': 20.0,  # 节点7
            'H': 20.0,  # 节点8
            'I': 20.0,  # 节点9
            'J': 20.0,  # 节点10
            'K': 15.0,  # 端/API/服务
            'L': 12.0,  # 冒烟结果
            'M': 15.0,  # 研发对应负责人
            'N': 20.0,  # showcase问题
            'O': 12.0,  # 是否核心功能
            'P': 12.0,  # 是否影响主流程
            'Q': 12.0   # 执行时间
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # 业务逻辑方法（复用之前的实现）
    def _determine_platform(self, test_case: Dict[str, Any]) -> str:
        """确定平台类型"""
        return 'Web/APP'  # 简化实现
    
    def _determine_smoke_result(self, test_case: Dict[str, Any]) -> str:
        """确定冒烟结果"""
        markers = test_case.get('markers', [])
        if 'flag-red' in markers:
            return '需关注'
        elif 'symbol-wrong' in markers:
            return '失败'
        else:
            return '通过'
    
    def _determine_developer(self, test_case: Dict[str, Any]) -> str:
        """确定负责人"""
        return '张三'  # 简化实现
    
    def _determine_showcase_issue(self, test_case: Dict[str, Any]) -> str:
        """确定showcase问题"""
        markers = test_case.get('markers', [])
        if 'flag-red' in markers:
            return '需重点关注'
        elif 'symbol-wrong' in markers:
            return '存在问题'
        else:
            return '无'
    
    def _create_summary_sheet(self, ws, metadata: Dict[str, Any], total_cases: int):
        """创建汇总表"""
        # 简化实现
        ws.cell(row=1, column=1, value='项目')
        ws.cell(row=1, column=2, value='值')
        
        summary_data = [
            ['源文件', metadata.get('source_file', '学习报告.xmind')],
            ['导出时间', datetime.now().strftime('%Y/%m/%d %H:%M:%S')],
            ['总用例数', total_cases],
            ['导出格式', '层级合并格式']
        ]
        
        for row, (key, value) in enumerate(summary_data, 2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)


def test_hierarchical_export():
    """测试层级导出功能"""
    logger.info("🧪 测试层级合并导出功能...")
    
    # 模拟测试数据
    test_data = {
        "smoke_test_suite": {
            "metadata": {
                "source_file": "学习报告.xmind",
                "export_time": datetime.now().isoformat(),
                "selected_markers": ["priority-1", "flag-red"],
                "total_cases": 4
            },
            "test_cases": [
                {
                    "test_path": "学习报告 > 展示规则 > 答题器参与课中不完成加油站验证 > 查看讲次详情",
                    "title": "不展示学习报告",
                    "markers": ["priority-1"]
                },
                {
                    "test_path": "学习报告 > 展示规则 > 答题器参与课中不完成加油站验证 > 其他操作",
                    "title": "其他测试",
                    "markers": ["priority-1"]
                },
                {
                    "test_path": "学习报告 > 高光时刻 > 高光时刻新增排行榜",
                    "title": "查看排行榜",
                    "markers": ["priority-1"]
                },
                {
                    "test_path": "学习报告 > 高光时刻 > 学生只有排行榜高光时刻",
                    "title": "显示验证",
                    "markers": ["flag-red"]
                }
            ]
        }
    }
    
    exporter = HierarchicalExcelExporter()
    output_file = exporter.export_with_hierarchical_merge(test_data)
    
    logger.info(f"✅ 层级合并测试完成，文件已生成: {output_file}")
    return output_file


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    test_hierarchical_export() 