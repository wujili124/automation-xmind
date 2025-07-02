#!/usr/bin/env python3
"""
测试Excel导出功能
"""
import requests
import json
import base64
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# API基础URL
API_BASE_URL = "http://localhost:8000"

def test_excel_export_functionality():
    """测试Excel导出功能的数据准备"""
    print("🚀 开始测试Excel导出功能数据准备...")
    
    # 1. 检查API健康状态
    print("\n1️⃣ 检查API健康状态...")
    try:
        response = requests.get(f"{API_BASE_URL}/")
        print(f"✅ API状态: {response.json()}")
    except Exception as e:
        print(f"❌ API连接失败: {e}")
        return
    
    # 2. 查找测试用的XMind文件
    print("\n2️⃣ 查找测试用的XMind文件...")
    test_files = []
    for root, dirs, files in os.walk(".."):
        for file in files:
            if file.endswith('.xmind'):
                test_files.append(os.path.join(root, file))
    
    if not test_files:
        print("❌ 未找到测试用的XMind文件")
        return
    
    test_file = test_files[0]
    print(f"✅ 使用测试文件: {test_file}")
    
    # 3. 分析XMind文件
    print("\n3️⃣ 分析XMind文件...")
    try:
        with open(test_file, 'rb') as f:
            files = {'file': (os.path.basename(test_file), f, 'application/octet-stream')}
            response = requests.post(f"{API_BASE_URL}/api/analyze", files=files)
        
        if response.status_code != 200:
            print(f"❌ 分析失败: {response.text}")
            return
        
        analysis_data = response.json()
        print(f"✅ 分析完成: 发现 {len(analysis_data['markers_found'])} 种标识符")
    
    except Exception as e:
        print(f"❌ 分析过程出错: {e}")
        return
    
    # 4. 测试导出功能
    print("\n4️⃣ 测试导出功能...")
    if not analysis_data['markers_found']:
        print("❌ 没有找到标识符，无法测试导出")
        return
    
    # 选择所有标识符进行导出
    selected_markers = [marker['markerId'] for marker in analysis_data['markers_found']]
    
    try:
        export_request = {
            "selected_markers": selected_markers,
            "file_data": analysis_data['file_data']
        }
        
        response = requests.post(f"{API_BASE_URL}/api/export", json=export_request)
        
        if response.status_code != 200:
            print(f"❌ 导出失败: {response.text}")
            return
        
        export_result = response.json()
        test_cases = export_result['smoke_test_suite']['test_cases']
        
        print(f"✅ 导出成功: 生成 {len(test_cases)} 个冒烟用例")
        
        # 5. 创建Excel文件（模拟前端Excel导出功能）
        print("\n5️⃣ 创建Excel文件（模拟前端功能）...")
        create_excel_file(export_result)
        
        print("\n🎉 Excel导出功能测试完成!")
        return True
        
    except Exception as e:
        print(f"❌ 导出过程出错: {e}")
        return False

def create_excel_file(export_result):
    """创建Excel文件（模拟前端功能）"""
    try:
        # 创建工作簿
        wb = Workbook()
        
        # 创建测试用例列表工作表
        ws1 = wb.active
        ws1.title = "冒烟测试用例"
        
        # 设置表头
        headers = [
            '用例ID', '测试用例标题', '模块', '优先级', '标识符', 
            '测试路径', '步骤数', '测试步骤', '是否核心功能', 
            '是否影响主流程', '执行时间'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
        
        # 添加测试用例数据
        test_cases = export_result['smoke_test_suite']['test_cases']
        for row, test_case in enumerate(test_cases, 2):
            ws1.cell(row=row, column=1).value = test_case['case_id']
            ws1.cell(row=row, column=2).value = test_case['title']
            ws1.cell(row=row, column=3).value = test_case['module']
            ws1.cell(row=row, column=4).value = test_case['priority']
            ws1.cell(row=row, column=5).value = ', '.join(test_case['markers'])
            ws1.cell(row=row, column=6).value = test_case['test_path']
            ws1.cell(row=row, column=7).value = len(test_case['steps'])
            
            # 合并测试步骤
            steps_text = '\n'.join([
                f"{step['step']}. {step['action']} -> {step['expected']}"
                for step in test_case['steps']
            ])
            ws1.cell(row=row, column=8).value = steps_text
            
            ws1.cell(row=row, column=9).value = '是' if test_case['smoke_criteria']['is_core_function'] else '否'
            ws1.cell(row=row, column=10).value = '是' if test_case['smoke_criteria']['affects_main_flow'] else '否'
            ws1.cell(row=row, column=11).value = test_case['smoke_criteria']['execution_time']
        
        # 设置列宽
        column_widths = [12, 30, 15, 8, 20, 40, 8, 50, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width
        
        # 创建汇总信息工作表
        ws2 = wb.create_sheet("导出汇总")
        
        summary_data = [
            ['项目', '值'],
            ['源文件', export_result['smoke_test_suite']['metadata']['source_file']],
            ['导出时间', datetime.fromisoformat(export_result['smoke_test_suite']['metadata']['export_time']).strftime('%Y-%m-%d %H:%M:%S')],
            ['选中标识符', ', '.join(export_result['smoke_test_suite']['metadata']['selected_markers'])],
            ['总用例数', export_result['smoke_test_suite']['metadata']['total_cases']],
            ['', ''],
            ['标识符统计', ''],
        ]
        
        # 统计优先级分布
        priority_stats = {}
        for test_case in test_cases:
            priority = test_case['priority']
            priority_stats[priority] = priority_stats.get(priority, 0) + 1
        
        for priority, count in priority_stats.items():
            priority_name = {'P1': 'P1 (高优先级)', 'P2': 'P2 (中优先级)', 'P3': 'P3 (标准优先级)'}.get(priority, priority)
            summary_data.append([priority_name, count])
        
        summary_data.extend([
            ['', ''],
            ['冒烟测试统计', ''],
            ['核心功能用例', sum(1 for tc in test_cases if tc['smoke_criteria']['is_core_function'])],
            ['影响主流程用例', sum(1 for tc in test_cases if tc['smoke_criteria']['affects_main_flow'])]
        ])
        
        # 写入汇总数据
        for row, (key, value) in enumerate(summary_data, 1):
            ws2.cell(row=row, column=1).value = key
            ws2.cell(row=row, column=2).value = value
            if row == 1 or key in ['标识符统计', '冒烟测试统计']:
                ws2.cell(row=row, column=1).font = ws2.cell(row=row, column=1).font.copy(bold=True)
                ws2.cell(row=row, column=2).font = ws2.cell(row=row, column=2).font.copy(bold=True)
        
        # 设置汇总表列宽
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 30
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f"冒烟测试用例_{timestamp}.xlsx"
        
        # 保存文件
        wb.save(filename)
        
        print(f"✅ Excel文件已创建: {filename}")
        print(f"   📊 包含 {len(test_cases)} 个测试用例")
        print(f"   📝 包含2个工作表: '冒烟测试用例' 和 '导出汇总'")
        
        # 验证文件是否可以正常打开
        try:
            verify_wb = load_workbook(filename)
            print(f"✅ Excel文件验证成功，工作表: {verify_wb.sheetnames}")
        except Exception as e:
            print(f"❌ Excel文件验证失败: {e}")
        
    except Exception as e:
        print(f"❌ 创建Excel文件失败: {e}")

if __name__ == "__main__":
    success = test_excel_export_functionality()
    if success:
        print("\n✅ Excel导出功能测试通过")
    else:
        print("\n❌ Excel导出功能测试失败") 