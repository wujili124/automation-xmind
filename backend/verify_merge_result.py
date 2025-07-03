#!/usr/bin/env python3
"""
验证Excel文件的合并效果
详细分析生成的Excel文件结构，确认层级合并是否正确实现
"""

import logging
from openpyxl import load_workbook
import os
import glob
from datetime import datetime

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def verify_excel_merge_result():
    """验证Excel文件的合并结果"""
    
    print("🔍 验证Excel文件合并效果")
    print("=" * 80)
    
    # 查找最新的增强层级合并文件
    enhanced_files = glob.glob("增强层级合并_冒烟测试用例_*.xlsx")
    if not enhanced_files:
        print("❌ 未找到增强层级合并文件")
        return False
    
    latest_file = sorted(enhanced_files)[-1]
    print(f"📁 分析文件: {latest_file}")
    
    try:
        wb = load_workbook(latest_file)
        ws = wb.active
        
        print(f"\n📊 基本信息:")
        print(f"   工作表数量: {len(wb.sheetnames)}")
        print(f"   主工作表: {ws.title}")
        print(f"   数据行数: {ws.max_row}")
        print(f"   数据列数: {ws.max_column}")
        
        # 检查表头
        print(f"\n📋 表头信息:")
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
            print(f"   列{col}: {header}")
        
        # 检查合并单元格
        print(f"\n🔗 合并单元格分析:")
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"   总合并区域数: {len(merged_ranges)}")
        
        if merged_ranges:
            merge_by_column = {}
            for merge_range in merged_ranges:
                col_letter = merge_range.start_cell.column_string
                start_row = merge_range.start_cell.row
                end_row = merge_range.end_cell.row
                value = ws.cell(row=start_row, column=merge_range.start_cell.column).value
                
                if col_letter not in merge_by_column:
                    merge_by_column[col_letter] = []
                
                merge_by_column[col_letter].append({
                    'range': str(merge_range),
                    'rows': f"{start_row}-{end_row}",
                    'span': end_row - start_row + 1,
                    'value': value
                })
            
            for col_letter, merges in sorted(merge_by_column.items()):
                col_index = ord(col_letter) - ord('A')
                col_name = headers[col_index] if col_index < len(headers) else f"列{col_letter}"
                print(f"\n   {col_letter}列 ({col_name}) 合并情况:")
                for merge in merges:
                    print(f"      范围: {merge['range']} | 跨度: {merge['span']}行 | 值: '{merge['value']}'")
        else:
            print("   ❌ 未发现任何合并单元格！")
        
        # 检查数据样本
        print(f"\n📝 数据样本 (前10行):")
        for row in range(2, min(12, ws.max_row + 1)):  # 跳过表头
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):  # 只显示前5列
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"   行{row}: {' | '.join(row_data)}")
        
        # 检查层级结构
        print(f"\n🌳 层级结构分析:")
        hierarchy_analysis = analyze_hierarchy_structure(ws)
        
        wb.close()
        
        # 总结
        print(f"\n📈 验证结果总结:")
        if merged_ranges:
            print("   ✅ 发现合并单元格功能")
            print(f"   ✅ 共有 {len(merged_ranges)} 个合并区域")
            if hierarchy_analysis['has_hierarchy']:
                print("   ✅ 检测到层级结构")
            else:
                print("   ⚠️ 合并存在但层级结构不清晰")
        else:
            print("   ❌ 合并功能未生效")
            print("   ❌ 可能存在算法问题")
        
        return len(merged_ranges) > 0
        
    except Exception as e:
        print(f"❌ 文件分析失败: {str(e)}")
        return False

def analyze_hierarchy_structure(ws):
    """分析层级结构"""
    
    analysis = {
        'has_hierarchy': False,
        'node_levels': {},
        'unique_nodes': set()
    }
    
    # 分析节点1-5列的数据
    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):  # 节点1-5列
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and str(cell_value).strip():
                level = col
                if level not in analysis['node_levels']:
                    analysis['node_levels'][level] = set()
                analysis['node_levels'][level].add(str(cell_value).strip())
                analysis['unique_nodes'].add(str(cell_value).strip())
    
    # 判断是否有层级结构
    if len(analysis['node_levels']) > 1:
        analysis['has_hierarchy'] = True
    
    print(f"   各层级节点统计:")
    for level in sorted(analysis['node_levels'].keys()):
        nodes = analysis['node_levels'][level]
        print(f"      节点{level}: {len(nodes)} 个唯一值")
    
    return analysis

def test_real_api_call():
    """测试真实的API调用并验证结果"""
    
    print("\n" + "=" * 80)
    print("🧪 测试真实API调用")
    print("=" * 80)
    
    import requests
    import base64
    
    # 准备测试数据
    xmind_path = "../学习报告.xmind"
    if not os.path.exists(xmind_path):
        print("❌ 测试文件不存在")
        return False
    
    with open(xmind_path, 'rb') as f:
        file_content = f.read()
        file_data = base64.b64encode(file_content).decode('utf-8')
    
    # 调用增强层级合并API
    request_data = {
        "selected_markers": ["priority-1", "priority-2", "flag-red"],
        "file_data": file_data
    }
    
    try:
        response = requests.post(
            "http://localhost:8000/api/export-enhanced-hierarchical",
            json=request_data,
            timeout=30
        )
        
        if response.status_code == 200:
            response_data = response.json()
            
            # 保存文件并分析
            excel_content = base64.b64decode(response_data['file_data'])
            filename = f"api_test_result_{datetime.now().strftime('%H%M%S')}.xlsx"
            
            with open(filename, 'wb') as f:
                f.write(excel_content)
            
            print(f"✅ API调用成功，文件已保存: {filename}")
            print(f"   文件大小: {len(excel_content)} 字节")
            
            # 立即分析这个文件
            return verify_single_file(filename)
        else:
            print(f"❌ API调用失败: {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ API测试失败: {str(e)}")
        return False

def verify_single_file(filename):
    """验证单个文件"""
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        merged_count = len(ws.merged_cells.ranges)
        print(f"   合并区域数: {merged_count}")
        print(f"   数据行数: {ws.max_row - 1}")
        
        if merged_count > 0:
            print("   ✅ 合并功能正常")
            
            # 显示合并详情
            for i, merge_range in enumerate(list(ws.merged_cells.ranges)[:5]):  # 显示前5个
                start_cell = merge_range.start_cell
                value = ws.cell(row=start_cell.row, column=start_cell.column).value
                print(f"   合并{i+1}: {merge_range} = '{value}'")
            
            wb.close()
            return True
        else:
            print("   ❌ 未检测到合并功能")
            wb.close()
            return False
            
    except Exception as e:
        print(f"   ❌ 文件验证失败: {str(e)}")
        return False

def main():
    """主函数"""
    
    print("🎯 开始验证Excel合并效果...")
    
    # 1. 验证已存在的文件
    existing_result = verify_excel_merge_result()
    
    # 2. 测试实时API调用
    api_result = test_real_api_call()
    
    # 3. 总结
    print("\n" + "=" * 80)
    print("🏁 最终验证结果")
    print("=" * 80)
    
    if existing_result and api_result:
        print("🎉 验证成功！合并功能正常工作")
        print("✅ 层级合并算法已正确实现")
    elif existing_result or api_result:
        print("⚠️ 部分功能正常，但可能存在不一致")
        print("🔧 建议进一步检查算法逻辑")
    else:
        print("❌ 合并功能未正确实现")
        print("🔧 需要修复层级合并算法")
        
        # 提供调试建议
        print("\n🛠️ 调试建议:")
        print("1. 检查数据分组逻辑是否正确")
        print("2. 验证合并区域计算算法")
        print("3. 确认openpyxl合并操作是否生效")
        print("4. 检查是否有异常导致合并跳过")

if __name__ == "__main__":
    main() 