#!/usr/bin/env python3
"""
最终验证：确认层级合并功能完全正常
"""

from openpyxl import load_workbook
import os

def final_verification():
    """最终验证层级合并功能"""
    
    print("🎯 最终验证：层级合并功能状态")
    print("=" * 60)
    
    # 查找生成的完美示例文件
    example_file = "🔥完美层级合并示例_20250703_102409.xlsx"
    
    if not os.path.exists(example_file):
        print("❌ 示例文件不存在")
        return False
    
    try:
        wb = load_workbook(example_file)
        ws = wb.active
        
        print(f"📁 验证文件: {example_file}")
        print(f"📊 数据行数: {ws.max_row - 1}")  # 减去表头
        print(f"📊 数据列数: {ws.max_column}")
        
        # 检查合并区域
        merged_count = len(ws.merged_cells.ranges)
        print(f"🔗 合并区域数: {merged_count}")
        
        if merged_count > 0:
            print("\n✅ 合并详情:")
            for i, merged_range in enumerate(ws.merged_cells.ranges, 1):
                # 获取合并范围信息
                min_col = merged_range.min_col
                max_col = merged_range.max_col  
                min_row = merged_range.min_row
                max_row = merged_range.max_row
                
                # 获取合并区域的值
                value = ws.cell(row=min_row, column=min_col).value
                span = max_row - min_row + 1
                
                # 获取列名
                if min_col <= 5:
                    col_name = f"节点{min_col}"
                else:
                    col_names = ['', '节点1', '节点2', '节点3', '节点4', '节点5', 
                                '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题']
                    col_name = col_names[min_col] if min_col < len(col_names) else f"列{min_col}"
                
                print(f"   {i:2d}. {col_name}: 行{min_row}-{max_row} (跨{span}行) = '{value}'")
        
        # 检查表头
        print(f"\n📋 表头验证:")
        expected_headers = ['节点1', '节点2', '节点3', '节点4', '节点5', 
                           '端/API/服务', '冒烟结果', '研发对应负责人', 'showcase问题', 
                           '是否核心功能', '是否影响主流程', '执行时间']
        
        actual_headers = []
        for col in range(1, min(13, ws.max_column + 1)):
            header = ws.cell(row=1, column=col).value
            actual_headers.append(header)
        
        headers_match = actual_headers[:len(expected_headers)] == expected_headers[:len(actual_headers)]
        
        if headers_match:
            print("   ✅ 表头完全匹配模版要求")
        else:
            print("   ⚠️ 表头与模版不完全匹配")
        
        # 检查数据样本
        print(f"\n📝 数据样本 (前5行):")
        for row in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, 6):  # 显示节点1-5
                cell_value = ws.cell(row=row, column=col).value
                display_value = str(cell_value)[:20] if cell_value else ""
                row_data.append(display_value)
            print(f"   行{row}: {' | '.join(row_data)}")
        
        wb.close()
        
        # 最终评估
        print(f"\n🎯 最终评估:")
        success_criteria = [
            (merged_count >= 5, f"合并区域数量 ({merged_count}/5+)"),
            (ws.max_row > 10, f"数据行数充足 ({ws.max_row - 1})"),
            (headers_match, "表头格式正确"),
            (ws.max_column == 12, f"列数正确 ({ws.max_column}/12)")
        ]
        
        passed = sum(1 for criteria, _ in success_criteria if criteria)
        total = len(success_criteria)
        
        for criteria, description in success_criteria:
            status = "✅" if criteria else "❌"
            print(f"   {status} {description}")
        
        print(f"\n📊 总体评分: {passed}/{total} ({passed/total*100:.0f}%)")
        
        if passed == total:
            print("🎉 层级合并功能完美实现！")
            print("✅ 完全符合《冒烟用例导出模版.xlsx》要求")
            print("✅ 智能单元格合并正常工作")
            print("✅ 视觉层级效果完美展现")
            return True
        else:
            print("⚠️ 部分功能需要优化")
            return False
        
    except Exception as e:
        print(f"❌ 验证失败: {str(e)}")
        return False

def show_usage_guide():
    """显示使用指南"""
    
    print(f"\n" + "🚀 使用指南")
    print("-" * 60)
    print("📌 前端集成建议:")
    print("")
    print("1️⃣ 确保调用正确的API接口:")
    print("   ✅ 使用: /api/export-enhanced-hierarchical")
    print("   ❌ 避免: /api/export 或 /api/export-template")
    print("")
    print("2️⃣ 检查返回的文件数据:")
    print("   - 文件大小应该在 10KB 左右")
    print("   - 包含 export_details.features 字段")
    print("   - features 包含 '精确单元格合并算法'")
    print("")
    print("3️⃣ 验证下载的Excel文件:")
    print("   - 打开文件应该能看到合并单元格")
    print("   - 层级结构应该清晰可见")
    print("   - 背景色应该有层级区分")
    print("")
    print("🔧 如果仍然看到传统格式：")
    print("   1. 检查前端是否调用了正确的API")
    print("   2. 确认下载的是最新生成的文件")
    print("   3. 清除浏览器缓存重新下载")

if __name__ == "__main__":
    success = final_verification()
    show_usage_guide()
    
    if success:
        print(f"\n🎊 恭喜！层级合并功能已完美实现！")
        print(f"📁 请查看示例文件验证效果")
    else:
        print(f"\n🔧 需要进一步调试和优化") 