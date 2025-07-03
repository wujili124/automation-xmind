#!/usr/bin/env python3
"""
测试严格格式导出功能
确保完全匹配目标文件格式
"""

import requests
import json
import base64
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# API基础URL
API_BASE_URL = "http://localhost:8000"

def test_strict_format_export():
    """测试严格格式导出API功能"""
    logger.info("🚀 开始测试严格格式导出...")
    
    try:
        # 1. 准备测试数据
        logger.info("\n1️⃣ 准备测试数据...")
        xmind_path = "../学习报告.xmind"
        
        if not os.path.exists(xmind_path):
            logger.error(f"❌ XMind文件不存在: {xmind_path}")
            return False
        
        # 读取文件
        with open(xmind_path, 'rb') as f:
            file_content = f.read()
            file_base64 = base64.b64encode(file_content).decode('utf-8')
        
        logger.info(f"✅ 成功读取XMind文件: {len(file_content)} 字节")
        
        # 2. 调用模版格式导出API
        logger.info("\n2️⃣ 调用模版格式导出API...")
        
        request_data = {
            "selected_markers": ["priority-1", "priority-2", "flag-red"],
            "file_data": file_base64
        }
        
        response = requests.post(
            f"{API_BASE_URL}/api/export-template",
            json=request_data,
            timeout=30
        )
        
        if response.status_code != 200:
            logger.error(f"❌ API调用失败: {response.status_code}")
            logger.error(f"错误信息: {response.text}")
            return False
        
        logger.info("✅ API调用成功")
        
        # 3. 保存生成的Excel文件
        logger.info("\n3️⃣ 保存生成的Excel文件...")
        
        response_data = response.json()
        excel_base64 = response_data.get('file_data')
        
        if not excel_base64:
            logger.error("❌ 响应中没有Excel文件内容")
            logger.error(f"实际响应: {response_data}")
            return False
        
        # 解码并保存文件
        excel_content = base64.b64decode(excel_base64)
        new_filename = f"strict_format_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with open(new_filename, 'wb') as f:
            f.write(excel_content)
        
        logger.info(f"✅ Excel文件已保存: {new_filename}")
        logger.info(f"📊 用例总数: {response_data.get('total_cases', 0)}")
        
        # 4. 详细对比格式
        logger.info("\n4️⃣ 详细对比格式...")
        compare_format_with_target(new_filename)
        
        return True
        
    except Exception as e:
        logger.error(f"❌ 测试失败: {str(e)}")
        return False

def compare_format_with_target(new_file):
    """详细对比新生成文件与目标文件的格式"""
    target_file = "冒烟测试用例_模版格式_2025-07-02_23-19-01.xlsx"
    
    logger.info(f"🔍 对比文件格式:")
    logger.info(f"   目标文件: {target_file}")
    logger.info(f"   新文件:   {new_file}")
    
    if not os.path.exists(target_file):
        logger.warning(f"⚠️ 目标文件不存在: {target_file}")
        return
    
    try:
        # 分析目标文件
        target_wb = load_workbook(target_file)
        new_wb = load_workbook(new_file)
        
        logger.info(f"\n📊 工作表对比:")
        logger.info(f"   目标文件工作表: {target_wb.sheetnames}")
        logger.info(f"   新文件工作表:   {new_wb.sheetnames}")
        
        if target_wb.sheetnames != new_wb.sheetnames:
            logger.error("❌ 工作表名称不匹配")
        else:
            logger.info("✅ 工作表名称匹配")
        
        # 对比每个工作表
        for sheet_name in target_wb.sheetnames:
            if sheet_name in new_wb.sheetnames:
                logger.info(f"\n📋 分析工作表: {sheet_name}")
                
                target_ws = target_wb[sheet_name]
                new_ws = new_wb[sheet_name]
                
                # 对比表头
                target_headers = []
                new_headers = []
                
                max_col = max(target_ws.max_column, new_ws.max_column)
                
                for col in range(1, max_col + 1):
                    target_val = target_ws.cell(row=1, column=col).value
                    new_val = new_ws.cell(row=1, column=col).value
                    target_headers.append(target_val)
                    new_headers.append(new_val)
                
                logger.info(f"   目标表头: {target_headers}")
                logger.info(f"   新文件表头: {new_headers}")
                
                if target_headers == new_headers:
                    logger.info("   ✅ 表头完全匹配")
                else:
                    logger.error("   ❌ 表头不匹配")
                    
                    # 显示差异
                    for i, (target_h, new_h) in enumerate(zip(target_headers, new_headers)):
                        if target_h != new_h:
                            logger.error(f"     列{i+1}: 目标='{target_h}' vs 新='{new_h}'")
                
                # 对比列宽
                logger.info(f"   📏 列宽对比:")
                width_match = True
                for col in range(1, max_col + 1):
                    col_letter = target_ws.cell(row=1, column=col).column_letter
                    target_width = target_ws.column_dimensions[col_letter].width
                    new_width = new_ws.column_dimensions[col_letter].width
                    
                    if abs(target_width - new_width) > 0.1:  # 允许小的浮点误差
                        logger.warning(f"     列{col_letter}: 目标={target_width:.1f} vs 新={new_width:.1f}")
                        width_match = False
                
                if width_match:
                    logger.info("   ✅ 列宽匹配")
                else:
                    logger.warning("   ⚠️ 列宽有差异")
                
                # 对比数据行数
                logger.info(f"   📊 数据行数: 目标={target_ws.max_row-1} vs 新={new_ws.max_row-1}")
                
                # 显示前几行数据作为样例
                logger.info(f"   📝 数据样例对比 (前3行):")
                for row in range(1, min(4, target_ws.max_row + 1)):
                    target_row = []
                    new_row = []
                    
                    for col in range(1, max_col + 1):
                        target_val = target_ws.cell(row=row, column=col).value
                        new_val = new_ws.cell(row=row, column=col).value
                        target_row.append(target_val)
                        new_row.append(new_val)
                    
                    logger.info(f"     第{row}行:")
                    logger.info(f"       目标: {target_row}")
                    logger.info(f"       新文件: {new_row}")
        
        target_wb.close()
        new_wb.close()
        
        logger.info("\n🎯 格式对比完成！")
        
    except Exception as e:
        logger.error(f"❌ 格式对比失败: {str(e)}")

def main():
    """主函数"""
    logger.info("🚀 开始严格格式导出测试...")
    
    success = test_strict_format_export()
    
    if success:
        logger.info("\n✅ 严格格式导出测试完成！")
        logger.info("🎯 请检查生成的Excel文件是否完全匹配目标格式")
    else:
        logger.error("\n❌ 严格格式导出测试失败！")
    
    return success

if __name__ == "__main__":
    main() 